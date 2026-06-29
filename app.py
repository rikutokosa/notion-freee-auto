"""
Notion → freee 自動仕訳登録システム
Flask Webアプリ本体

エンドポイント:
  GET  /                     ダッシュボード
  GET  /preview              経理対応待ちプレビュー
  GET  /log                  処理ログ
  GET  /chat                 自然言語指示チャット
  POST /chat                 チャット送信
  GET  /auth/freee           freee OAuth認証開始
  GET  /auth/freee/callback  freee OAuth認証コールバック
  POST /run                  手動で全件処理
  POST /run/single           1件処理
  POST /polling/start        ポーリング開始
  POST /polling/stop         ポーリング停止
  GET  /api/pending          経理対応待ち一覧（JSON）
  GET  /api/logs             処理ログ（JSON）
  GET  /api/status           システム状態（JSON）
  POST /api/refresh_cache    freeeマスタキャッシュ更新
"""
import os
import json
import logging
import threading
from datetime import datetime
import tempfile
from pathlib import Path
import sqlite3

# ============================================================
# 会話履歴DB（ボリューム永続化）
# ============================================================
_VOLUME_PATH = os.environ.get("RAILWAY_VOLUME_MOUNT_PATH", "/data")
_DB_PATH = os.path.join(_VOLUME_PATH, "chat_history.db")

def _get_db():
    """SQLite接続を返す。DBファイルが存在しない場合は作成する。"""
    os.makedirs(os.path.dirname(_DB_PATH), exist_ok=True)
    conn = sqlite3.connect(_DB_PATH)
    conn.execute("""
        CREATE TABLE IF NOT EXISTS chat_sessions (
            session_id TEXT NOT NULL,
            seq        INTEGER NOT NULL,
            role       TEXT NOT NULL,
            content    TEXT NOT NULL,
            created_at TEXT NOT NULL DEFAULT (datetime('now')),
            PRIMARY KEY (session_id, seq)
        )
    """)
    # 運用ルール・メモテーブル
    conn.execute("""
        CREATE TABLE IF NOT EXISTS rules_notes (
            id         INTEGER PRIMARY KEY AUTOINCREMENT,
            category   TEXT NOT NULL,
            title      TEXT NOT NULL,
            content    TEXT NOT NULL,
            created_at TEXT NOT NULL DEFAULT (datetime('now','localtime')),
            updated_at TEXT NOT NULL DEFAULT (datetime('now','localtime'))
        )
    """)
    # 実行ログ永続保存テーブル（自動転記・請求書照合・総合振込）
    conn.execute("""
        CREATE TABLE IF NOT EXISTS execution_logs (
            id          INTEGER PRIMARY KEY AUTOINCREMENT,
            log_type    TEXT NOT NULL,
            executed_at TEXT NOT NULL,
            trigger     TEXT NOT NULL DEFAULT 'manual',
            summary     TEXT NOT NULL,
            detail      TEXT NOT NULL DEFAULT '{}',
            has_error   INTEGER NOT NULL DEFAULT 0
        )
    """)
    conn.commit()
    return conn


def _save_execution_log(log_type: str, summary: dict, detail: dict = None, trigger: str = 'manual', has_error: bool = False):
    """実行ログをSQLiteに永続保存する"""
    from datetime import datetime, timezone, timedelta
    JST = timezone(timedelta(hours=9))
    executed_at = datetime.now(JST).strftime("%Y-%m-%d %H:%M")
    conn = _get_db()
    try:
        conn.execute(
            "INSERT INTO execution_logs (log_type, executed_at, trigger, summary, detail, has_error) VALUES (?, ?, ?, ?, ?, ?)",
            (log_type, executed_at, trigger, json.dumps(summary, ensure_ascii=False), json.dumps(detail or {}, ensure_ascii=False), 1 if has_error else 0)
        )
        conn.commit()
    finally:
        conn.close()


def _get_execution_logs(log_type: str, limit: int = 50):
    """実行ログをSQLiteから取得する"""
    conn = _get_db()
    try:
        rows = conn.execute(
            "SELECT id, log_type, executed_at, trigger, summary, detail, has_error FROM execution_logs WHERE log_type=? ORDER BY id DESC LIMIT ?",
            (log_type, limit)
        ).fetchall()
        return [
            {
                "id": r[0], "log_type": r[1], "executed_at": r[2],
                "trigger": r[3],
                "summary": json.loads(r[4]),
                "detail": json.loads(r[5]),
                "has_error": bool(r[6])
            }
            for r in rows
        ]
    finally:
        conn.close()


from flask import Flask, render_template, request, redirect, jsonify, url_for, send_file

from freee_client import (
    get_auth_url, exchange_code_for_token, get_valid_token,
    load_token, is_token_expired, clear_master_cache, get_master_cache,
    create_deal, upload_receipt, create_partner,
    search_deals, search_invoices, delete_deal, delete_invoice,
    # エージェント拡張ツール
    get_deal, update_deal, list_deals, get_account_item_balance,
    list_invoices, register_invoice_agent,
    execute_delete_deal, execute_delete_invoice,
    resolve_partner_id,
)
from matcher import run_matching
from payment import build_fb_file
from notion_client import fetch_pending_records, get_record
from rules import build_journal_entries
from processor import run_once, process_single_by_id, processing_log

# ============================================================
# ロギング設定
# ============================================================
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(name)s: %(message)s",
)
logger = logging.getLogger(__name__)

# ============================================================
# Flask アプリ
# ============================================================
app = Flask(__name__)
app.secret_key = os.environ.get("SECRET_KEY", "notion-freee-secret-2026")
app.config['MAX_CONTENT_LENGTH'] = 20 * 1024 * 1024  # 20MB

POLL_INTERVAL = int(os.environ.get("POLL_INTERVAL", "3600"))

# ============================================================
# バックグラウンド自動転記
# ============================================================
_polling_thread = None
_polling_active = False

# 停止フラグ：Railway環境変数 FREEE_AUTO_STOPPED=1 で永続化
# ファイルはコンテナ再起動で消えるため使わない
# メモリ内フラグ（プロセス内永続）
_manually_stopped: bool = os.environ.get("FREEE_AUTO_STOPPED", "0") == "1"
# 仕訳アシスタント登録ログ（メモリ内、最新200件）
assistant_log: list = []


def _is_manually_stopped() -> bool:
    """手動停止中かどうかを返す（メモリ内フラグ）"""
    return _manually_stopped


def _set_manually_stopped(stopped: bool):
    """停止フラグをメモリ内に設定する"""
    global _manually_stopped
    _manually_stopped = stopped
    # Railway環境変数にも書き込む（プロセス内のみ有効、再起動後は環境変数の初期値に戻る）
    os.environ["FREEE_AUTO_STOPPED"] = "1" if stopped else "0"


def start_polling(force: bool = False):
    """自動転記を開始する。force=Trueの場合は手動停止フラグを無視して開始。"""
    global _polling_thread, _polling_active
    # 手動停止中は開始しない（forceの場合のみ例外）
    if not force and _is_manually_stopped():
        logger.info("自動転記: 手動停止中のため自動開始をスキップ")
        return
    if _polling_thread and _polling_thread.is_alive():
        return
    _polling_active = True
    _set_manually_stopped(False)

    def loop():
        import time
        while _polling_active:
            try:
                token = load_token()
                if token and not is_token_expired(token):
                    logger.info("freee自動転記実行")
                    run_once(db_type="all")
                else:
                    logger.warning("freeeトークン未設定または期限切れ。freee自動転記をスキップ。")
            except Exception as e:
                logger.exception(f"freee自動転記エラー: {e}")
            time.sleep(POLL_INTERVAL)

    _polling_thread = threading.Thread(target=loop, daemon=True)
    _polling_thread.start()
    logger.info(f"freee自動転記開始 (間隔: {POLL_INTERVAL}秒)")


# ============================================================
# ページルート
# ============================================================
@app.route("/")
def index():
    """ダッシュボード"""
    token_ok = False
    try:
        get_valid_token()
        token_ok = True
    except Exception:
        pass

    return render_template(
        "index.html",
        token_ok=token_ok,
        logs=processing_log[:50],
    )


@app.route("/preview")
def preview():
    """②経理対応待ちレコードのプレビュー"""
    db_type = request.args.get("db", "all")
    try:
        records = fetch_pending_records(db_type)
        previews = []
        for record in records:
            journal = build_journal_entries(record)
            previews.append({
                "page_id": record.get("id"),
                "phase": journal.get("phase", ""),
                "job_db": journal.get("job_db", ""),
                "nyusha_date": journal.get("nyusha_date", ""),
                "action": journal.get("action", ""),
                "message": journal.get("message", ""),
                "needs_invoice": journal.get("needs_invoice", False),
                "original_status": journal.get("original_status", ""),
                "db_type": record.get("_db_type", "honten"),
                "sales_entry": journal.get("sales_entry"),
                "purchase_entry": journal.get("purchase_entry"),
                "pca_entry": journal.get("pca_entry"),
            })
        return render_template("preview.html", previews=previews, db_type=db_type, assistant_log=assistant_log[:50])
    except Exception as e:
        logger.exception("プレビューエラー")
        return render_template("error.html", message=str(e))


@app.route("/log")
def log_page():
    """処理ログ一覧"""
    return render_template("log.html", logs=processing_log)


@app.route("/chat")
def chat():
    """旧チャット指示ページ → ホームにリダイレクト"""
    return redirect(url_for("index"))


# ============================================================
# freee OAuth認証
# ============================================================
@app.route("/auth/freee")
def auth_freee():
    """freee OAuth認証開始"""
    redirect_uri = url_for("auth_freee_callback", _external=True)
    # APP_BASE_URL が設定されている場合はそちらを優先
    base_url = os.environ.get("APP_BASE_URL", "").rstrip("/")
    if base_url:
        redirect_uri = f"{base_url}/auth/freee/callback"
    auth_url = get_auth_url(redirect_uri, state="notion_freee")
    return redirect(auth_url)


@app.route("/auth/freee/callback")
def auth_freee_callback():
    """freee OAuth認証コールバック"""
    code = request.args.get("code")
    error = request.args.get("error")

    if error:
        return render_template("error.html", message=f"認証エラー: {error}")
    if not code:
        return render_template("error.html", message="認証コードが取得できませんでした")

    try:
        redirect_uri = url_for("auth_freee_callback", _external=True)
        base_url = os.environ.get("APP_BASE_URL", "").rstrip("/")
        if base_url:
            redirect_uri = f"{base_url}/auth/freee/callback"
        exchange_code_for_token(code, redirect_uri)
        return redirect(url_for("index"))
    except Exception as e:
        logger.exception("トークン取得エラー")
        return render_template("error.html", message=f"トークン取得エラー: {str(e)}")


@app.route("/api/get_refresh_token")
def api_get_refresh_token():
    """現在のリフレッシュトークンを返す（Railway環境変数設定用）"""
    token_data = load_token()
    if not token_data:
        return jsonify({"error": "トークン未認証。/auth/freee から認証してください。"}), 401
    refresh_token = token_data.get("refresh_token", "")
    if not refresh_token:
        return jsonify({"error": "リフレッシュトークンがありません。再認証してください。"}), 401
    return jsonify({
        "refresh_token": refresh_token,
        "instruction": "Railwayダッシュボードの環境変数に FREEE_REFRESH_TOKEN = <上記の値> を設定してください。以後はデプロイ後も自動復元されます。"
    })


# ============================================================
# 手動実行
# ============================================================
@app.route("/run", methods=["POST"])
def run_manual():
    """手動で全件処理を実行する"""
    data = request.get_json() or {}
    db_type = data.get("db_type", "all")
    dry_run = data.get("dry_run", False)
    try:
        results = run_once(db_type=db_type, dry_run=dry_run)
        success = sum(1 for r in results if r.get("status") == "success")
        errors = sum(1 for r in results if r.get("status") == "error")
        reviews = sum(1 for r in results if r.get("status") == "review")
        skips = sum(1 for r in results if r.get("status") == "skip")
        if not dry_run:
            _save_execution_log(
                log_type="auto_transfer",
                summary={"total": len(results), "success": success, "errors": errors, "reviews": reviews, "skips": skips, "db_type": db_type},
                detail={"results": [{"name": r.get("name",""), "status": r.get("status",""), "message": r.get("message",""), "action": r.get("action",""), "sales_id": r.get("sales_id"), "purchase_id": r.get("purchase_id"), "pca_id": r.get("pca_id"), "invoice_id": r.get("invoice_id")} for r in results]},
                trigger="manual",
                has_error=errors > 0
            )
        return jsonify({
            "status": "ok",
            "total": len(results),
            "success": success,
            "errors": errors,
            "reviews": reviews,
            "dry_run": dry_run,
            "results": results,
        })
    except Exception as e:
        logger.exception("手動実行エラー")
        _save_execution_log(log_type="auto_transfer", summary={"total": 0, "success": 0, "errors": 1, "error_message": str(e)}, trigger="manual", has_error=True)
        return jsonify({"status": "error", "message": str(e)}), 500


@app.route("/run/single", methods=["POST"])
def run_single():
    """1件を処理する"""
    data = request.get_json() or {}
    page_id = data.get("page_id")
    db_type = data.get("db_type", "honten")
    dry_run = data.get("dry_run", False)
    if not page_id:
        return jsonify({"status": "error", "message": "page_id が必要です"}), 400
    try:
        result = process_single_by_id(page_id, db_type=db_type, dry_run=dry_run)
        return jsonify(result)
    except Exception as e:
        logger.exception("単件処理エラー")
        return jsonify({"status": "error", "message": str(e)}), 500


@app.route("/polling/start", methods=["POST"])
def polling_start():
    """freee自動転記を手動で開始（手動停止フラグをクリアして強制開始）"""
    start_polling(force=True)
    return jsonify({"status": "ok", "message": "freee自動転記開始"})


@app.route("/polling/stop", methods=["POST"])
def polling_stop():
    """freee自動転記を手動停止（フラグをファイルに保存して再認証後も自動開始しない）"""
    global _polling_active
    _polling_active = False
    _set_manually_stopped(True)
    return jsonify({"status": "ok", "message": "freee自動転記を停止しました。手動で「自動転記開始」を押すまで停止します。"})


# ============================================================
# チャット指示（自然言語）
# ============================================================
def handle_chat_command(message: str) -> str:
    """自然言語コマンドを解釈して実行する"""
    msg = message.lower()

    # 全件実行
    if any(k in msg for k in ["実行", "登録", "転記", "処理", "run", "今すぐ", "全件"]):
        try:
            results = run_once(db_type="all")
            success = [r for r in results if r.get("status") == "success"]
            errors = [r for r in results if r.get("status") == "error"]
            reviews = [r for r in results if r.get("status") == "review"]
            lines = [f"処理完了: {len(results)}件"]
            if success:
                lines.append(f"✅ 成功: {len(success)}件")
                for r in success:
                    lines.append(f"  - {r.get('phase', '')} [{r.get('action','')}]: {r.get('message','')}")
            if reviews:
                lines.append(f"⚠️ 要確認: {len(reviews)}件")
                for r in reviews:
                    lines.append(f"  - {r.get('phase', '')}: {r.get('message', '')}")
            if errors:
                lines.append(f"❌ エラー: {len(errors)}件")
                for r in errors:
                    lines.append(f"  - {r.get('phase', '')}: {r.get('message', '')}")
            if not results:
                lines.append("経理対応待ちのレコードはありませんでした。")
            return "\n".join(lines)
        except Exception as e:
            return f"エラーが発生しました: {str(e)}"

    # 件数確認
    if any(k in msg for k in ["状態", "ステータス", "確認", "status", "何件", "いくつ"]):
        try:
            records = fetch_pending_records("all")
            if not records:
                return "経理対応待ちのレコードはありません。"
            lines = [f"経理対応待ちレコード: {len(records)}件"]
            for r in records:
                j = build_journal_entries(r)
                db = r.get("_db_type", "honten")
                lines.append(f"  [{db}] {j.get('phase','')} ({j.get('original_status','')}) → {j.get('action','')}")
            return "\n".join(lines)
        except Exception as e:
            return f"Notion接続エラー: {str(e)}"

    # ログ確認
    if any(k in msg for k in ["ログ", "履歴", "log", "history"]):
        if not processing_log:
            return "処理履歴はまだありません。"
        lines = ["直近の処理履歴:"]
        icons = {"success": "✅", "error": "❌", "review": "⚠️", "dry_run": "👁"}
        for r in processing_log[:10]:
            icon = icons.get(r.get("status", ""), "?")
            ts = r.get("timestamp", "")[:16]
            lines.append(f"{icon} {ts} {r.get('phase','')} - {r.get('message','')}")
        return "\n".join(lines)

    # 認証
    if any(k in msg for k in ["認証", "ログイン", "auth", "token"]):
        return "freee認証は /auth/freee から行ってください。"

    # ポーリング制御
    if any(k in msg for k in ["停止", "stop", "止め"]):
        global _polling_active
        _polling_active = False
        return "自動ポーリングを停止しました。"
    if any(k in msg for k in ["開始", "start", "再開"]):
        start_polling()
        return "自動ポーリングを開始しました。"

    # プレビュー
    if any(k in msg for k in ["プレビュー", "preview", "確認", "見せ"]):
        return "プレビューは /preview から確認できます。"

    return (
        "以下のコマンドが使えます:\n"
        "・「今すぐ処理して」「全件登録」 → 経理対応待ちレコードをfreeeに登録\n"
        "・「何件ある？」「状態確認」 → 未処理件数と内容を確認\n"
        "・「ログ見せて」「履歴」 → 処理履歴を表示\n"
        "・「ポーリング開始/停止」 → 自動処理の制御\n"
        "・「プレビュー」 → 登録前の仕訳内容を確認\n"
    )


# ============================================================
# API エンドポイント（JSON）
# ============================================================
@app.route("/api/pending")
def api_pending():
    db_type = request.args.get("db", "all")
    try:
        records = fetch_pending_records(db_type)
        result = []
        for r in records:
            j = build_journal_entries(r)
            result.append({
                "page_id": r.get("id"),
                "phase": j.get("phase"),
                "job_db": j.get("job_db"),
                "nyusha_date": j.get("nyusha_date"),
                "action": j.get("action"),
                "message": j.get("message"),
                "original_status": j.get("original_status"),
                "db_type": r.get("_db_type", "honten"),
            })
        return jsonify({"count": len(result), "records": result})
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/api/logs")
def api_logs():
    limit = int(request.args.get("limit", 100))
    return jsonify({"logs": processing_log[:limit]})


@app.route("/api/execution_logs")
def api_execution_logs():
    """実行ログを取得する。log_typeパラメータで種別を指定。"""
    log_type = request.args.get("log_type", "auto_transfer")
    limit = int(request.args.get("limit", 50))
    logs = _get_execution_logs(log_type, limit)
    return jsonify({"logs": logs, "log_type": log_type})


@app.route("/api/status")
def api_status():
    token_ok = False
    try:
        get_valid_token()
        token_ok = True
    except Exception:
        pass
    import os
    openai_key = os.environ.get("OPENAI_API_KEY", "")
    return jsonify({
        "token_ok": token_ok,
        "freee_connected": token_ok,
        "mode": "manual",
        "log_count": len(processing_log),
        "process_count": len(processing_log),
        "openai_key_set": bool(openai_key),
        "openai_key_prefix": openai_key[:10] + "..." if openai_key else "",
    })


@app.route("/api/refresh_cache", methods=["POST"])
def api_refresh_cache():
    clear_master_cache()
    return jsonify({"message": "マスタキャッシュをクリアしました"})


# ============================================================
# 自動転記プレビューAPI
# ============================================================
@app.route("/api/preview")
def api_preview():
    """経理対応待ちレコードの仕訳プレビューをJSONで返す（自動転記タブ用）"""
    db_type = request.args.get("db", "all")
    try:
        records = fetch_pending_records(db_type)
        previews = []
        for record in records:
            journal = build_journal_entries(record)
            if journal.get("action") == "skip":
                continue
            previews.append({
                "page_id": record.get("id"),
                "phase": journal.get("phase", ""),
                "job_db": journal.get("job_db", ""),
                "nyusha_date": journal.get("nyusha_date", ""),
                "action": journal.get("action", ""),
                "message": journal.get("message", ""),
                "needs_invoice": journal.get("needs_invoice", False),
                "original_status": journal.get("original_status", ""),
                "db_type": record.get("_db_type", "honten"),
                "sales_entry": journal.get("sales_entry"),
                "purchase_entry": journal.get("purchase_entry"),
                "pca_entry": journal.get("pca_entry"),
            })
        return jsonify({"previews": previews})
    except Exception as e:
        logger.exception("プレビューAPIエラー")
        return jsonify({"error": str(e)}), 500


# ============================================================
# 書類照合
# ============================================================
@app.route("/api/match/preview", methods=["GET"])
def api_match_preview():
    """書類照合プレビュー（dry_run=True）"""
    try:
        result = run_matching(dry_run=True)
        return jsonify(result)
    except Exception as e:
        logger.exception("書類照合プレビューエラー")
        return jsonify({"error": str(e)}), 500


@app.route("/api/match/execute", methods=["POST"])
def api_match_execute():
    """書類照合実行（dry_run=False）"""
    try:
        result = run_matching(dry_run=False)
        _save_execution_log(
            log_type="invoice_match",
            summary={
                "total_receipts": result.get("total_receipts", 0),
                "matched_count": result.get("matched_count", 0),
                "unmatched_count": result.get("unmatched_count", 0),
                "ai_ocr_count": result.get("ai_ocr_count", 0),
            },
            detail={
                "matched": result.get("matched", []),
                "unmatched": result.get("unmatched", []),
                "errors": result.get("errors", []),
            },
            trigger="manual",
            has_error=len(result.get("errors", [])) > 0
        )
        return jsonify(result)
    except Exception as e:
        logger.exception("書類照合実行エラー")
        _save_execution_log(log_type="invoice_match", summary={"total_receipts": 0, "matched_count": 0, "unmatched_count": 0, "error_message": str(e)}, trigger="manual", has_error=True)
        return jsonify({"error": str(e)}), 500


@app.route("/api/freee_master")
def api_freee_master():
    """freeeマスタデータ（部門・メモタグ・勘定科目・取引先）を返す（仕訳アシスタント用）"""
    try:
        cache = get_master_cache()
        return jsonify({
            "sections": [{"id": s.get("id"), "name": s.get("name")} for s in cache.get("sections", [])],
            "tags": [{"id": t.get("id"), "name": t.get("name")} for t in cache.get("tags", [])],
            "account_items": [{"id": a.get("id"), "name": a.get("name")} for a in cache.get("account_items", [])],
            "partners": [{"id": p.get("id"), "name": p.get("name")} for p in cache.get("partners", [])],
        })
    except Exception as e:
        return jsonify({"error": str(e)}), 500


# ============================================================
# 仕訳アシスタント
# ============================================================
@app.route("/assistant")
def assistant():
    """旧仕訳アシスタントページ → ホームにリダイレクト"""
    return redirect(url_for("index"))


@app.route("/api/assistant/register", methods=["POST"])
def api_assistant_register():
    """フォーム入力から直接freeeに仕訳登録する"""
    data = request.get_json() or {}
    try:
        cache = get_master_cache()
        deal_type = data.get("deal_type", "income")
        issue_date = data["issue_date"]
        due_date = data.get("due_date")
        partner_name = data.get("partner_name")
        details_raw = data.get("details", [])
        repeat = data.get("repeat", False)
        repeat_count = int(data.get("repeat_count", 1))

        # 明細行に section_name を付与（detail内にあれば使う）
        details = []
        section_name = None
        for d in details_raw:
            sn = d.pop("section_name", None)
            if sn and not section_name:
                section_name = sn
            details.append(d)

        registered_ids = []
        from dateutil.relativedelta import relativedelta
        from datetime import date
        base_date = date.fromisoformat(issue_date)

        count = repeat_count if repeat else 1
        for i in range(count):
            current_date = base_date + relativedelta(months=i)
            current_due = None
            if due_date:
                base_due = date.fromisoformat(due_date)
                current_due = (base_due + relativedelta(months=i)).isoformat()

            entry = {
                "issue_date": current_date.isoformat(),
                "due_date": current_due,
                "partner_name": partner_name,
                "section_name": section_name,
                "details": details,
            }
            deal = create_deal(entry, deal_type, cache)
            registered_ids.append(deal.get("id"))

        # assistant_logに記録
        from datetime import datetime as _dt
        for i, rid in enumerate(registered_ids):
            assistant_log.insert(0, {
                "source": "assistant",
                "freee_id": rid,
                "issue_date": (base_date + __import__('dateutil.relativedelta', fromlist=['relativedelta']).relativedelta(months=i)).isoformat() if repeat else issue_date,
                "partner_name": partner_name or "",
                "deal_type": deal_type,
                "amount": sum(d.get("amount", 0) for d in details),
                "registered_at": _dt.now().strftime("%Y-%m-%d %H:%M"),
                "note": "",
            })
        del assistant_log[200:]
        deal_urls = [f"https://secure.freee.co.jp/deals#deal_id={did}" for did in registered_ids]
        return jsonify({"status": "ok", "registered": len(registered_ids), "ids": registered_ids, "urls": deal_urls})
    except Exception as e:
        logger.exception("仕訳アシスタント登録エラー")
        return jsonify({"status": "error", "message": str(e)}), 500


@app.route("/api/assistant/register_bulk", methods=["POST"])
def api_assistant_register_bulk():
    """AI生成の仕訳エントリを一括登録する"""
    data = request.get_json() or {}
    entries = data.get("entries", [])
    try:
        cache = get_master_cache()
        registered_ids = []
        for entry in entries:
            deal_type = entry.pop("deal_type", "income")
            deal = create_deal(entry, deal_type, cache)
            registered_ids.append(deal.get("id"))
        # assistant_logに記録
        from datetime import datetime as _dt2
        for entry_orig, rid in zip(entries, registered_ids):
            assistant_log.insert(0, {
                "source": "assistant",
                "freee_id": rid,
                "issue_date": entry_orig.get("issue_date", ""),
                "partner_name": entry_orig.get("partner_name", ""),
                "deal_type": entry_orig.get("deal_type", "income"),
                "amount": sum(d.get("amount", 0) for d in entry_orig.get("details", [])),
                "registered_at": _dt2.now().strftime("%Y-%m-%d %H:%M"),
                "note": "",
            })
        del assistant_log[200:]
        deal_urls = [f"https://secure.freee.co.jp/deals#deal_id={did}" for did in registered_ids]
        return jsonify({"status": "ok", "registered": len(registered_ids), "ids": registered_ids, "urls": deal_urls})
    except Exception as e:
        logger.exception("AI仕訳一括登録エラー")
        return jsonify({"status": "error", "message": str(e)}), 500


@app.route("/api/assistant/ai", methods=["POST"])
def api_assistant_ai():
    """自然言語指示をFunction Calling対応のAIエージェントで処理する"""
    import os, json as _json, re
    from datetime import date as _date
    data = request.get_json() or {}
    user_message = data.get("message", "")
    history = data.get("history", [])
    master = data.get("master", {})

    try:
        import requests as req
        openai_key = os.environ.get("OPENAI_API_KEY", "")
        openai_base = os.environ.get("OPENAI_API_BASE", "https://api.openai.com/v1")
        if not openai_key:
            return jsonify({"error": "OpenAI APIキーが設定されていません"}), 500

        # 現在日付・年度情報
        today = _date.today()
        today_str = today.strftime("%Y年%m月%d日")
        fiscal_year = today.year if today.month >= 4 else today.year - 1
        fiscal_start = f"{fiscal_year}-04-01"
        fiscal_end = f"{fiscal_year + 1}-03-31"
        last_month = today.month - 1 if today.month > 1 else 12
        last_month_year = today.year if today.month > 1 else today.year - 1

        # マスタデータ：サーバー側で最新を取得
        try:
            server_master = get_master_cache()
        except Exception:
            server_master = master

        def _names(lst): return ", ".join([x["name"] for x in lst if x.get("name")])
        # 部門は「親部門名：子部門名」形式で表示し、子部門のみ使用可能であることを明示する
        raw_sections = server_master.get("sections", [])
        section_id_map = {s["id"]: s["name"] for s in raw_sections}
        sections_display = []
        for s in raw_sections:
            if s.get("parent_id"):
                parent_name = section_id_map.get(s["parent_id"], "")
                sections_display.append(f"{parent_name}：{s['name']}" if parent_name else s["name"])
        sections_list = ", ".join(sections_display) if sections_display else _names(raw_sections)
        tags_list = _names(server_master.get("tags", []))
        account_items_list = _names(server_master.get("account_items", []))
        partners_list = _names(server_master.get("partners", []))

        system_prompt = f"""あなたはfreee会計を自律的に操作できるAIエージェントです。
ユーザーの自然言語による指示を理解し、必要なツールを組み合わせて目的を達成してください。

# 基本情報
- 今日の日付: {today_str}
- 今年度: {fiscal_year}年4月〜{fiscal_year+1}年3月
- 先月: {last_month_year}年{last_month}月

# freeeマスタデータ
取引先: {partners_list}
部門（「親部門：子部門」形式）: {sections_list}
※ 同名の子部門が複数ある場合（例:「その他」が本店とCSSの両方にある）は、必ず「親部門：子部門」形式で指定すること。例: 「本店：その他」「CSS：その他」
メモタグ: {tags_list}
勘定科目: {account_items_list}

# エージェントとしての行動原則

## 1. 自律的な判断と実行
- ユーザーの意図を汲み取り、必要なツールを自分で選択して実行する
- 不明な点は推測して進める（ユーザーへの質問より行動を優先）
- 複数ステップが必要な場合は、ツールを連続して呼び出して完結させる
- 「検索→確認→登録」「検索→集計→報告」など複合的な処理も自律的にこなす

## 2. 登録・更新処理
- 取引先名・勘定科目名はマスタ一覧から最も近いものを推測して使用する
  例: 「ステリファイ」→「stellify」、「売上」→「売上高」、「広告費」→「広告宣伝費」
- **ユーザーが勘定科目を明示的に指定した場合は、必ずその科目を使用する**。AIが「より適切」と判断しても勝手に変更しない。
  例: ユーザーが「前払費用で」と言ったら「前払費用」を使う（「地代家賃」に変えない）
- 取引先がマスタに存在しない場合はcreate_partnerで自動作成してから登録する
- 日付の解釈: 「先月」→{last_month_year}-{last_month:02d}、「今月」→{today.year}-{today.month:02d}、「7月」→{fiscal_year}-07
- 複数件の登録指示（スケジュール表・明細表など）は全件を処理する

## 2-A. 仕訳登録の必須項目（register_deal呼び出し前に必ず確認）
仕訳を登録する際は、以下の4項目が**すべて揃っている**ことを確認してからregister_dealを呼び出すこと。
1. **発生日**（issue_date）: YYYY-MM-DD形式の日付
2. **取引先**（partner_name）: freeeマスタに登録されている取引先名
3. **決済期日**（due_date）: YYYY-MM-DD形式の日付
4. **対応部門**（section_name）: freeeマスタに登録されている部門名

**いずれか1つでも欠けている場合は、register_dealを呼び出さず、不足している項目を具体的に列挙してユーザーに聞き返すこと。**
例: 「登録に必要な情報が不足しています。以下を教えてください：\n- 決済期日（例: 2026-07-31）\n- 対応部門（例: 本店CA）」
全項目が揃ったことを確認してからregister_dealを呼び出すこと。

## 2-B. 仕訳更新の手順（update_deal呼び出し前に必ず実行）
仕訳を更新する際は必ず以下の順序で実行すること。
1. **search_dealsで対象仕訳を検索**：取引先名・日付範囲で検索し、実際のIDを取得する
2. **get_dealで対象仕訳の詳細を取得**：現在の明細行（勘定科目・金額・部門・摘要）を必ず確認する
3. **即座にupdate_dealを呼び出す**：テキストで「よろしいですか？」と確認を求めない。フロントが確認ボタンを表示するので、AIは即座にツールを呼び出すこと。

更新時の重要ルール:
- **既存仕訳の勘定科目構成を維持する**：ユーザーが明示的に科目変更を指示しない限り、各明細行の勘定科目は現在のまま維持する。例えば「前払費用」と「地代家賃」の2行がある仕訳で「金額を修正して」と言われたら、科目はそのままで金額だけ変える。画像や資料に別の科目名が書いてあっても、それは内訳や参考情報であり、既存仕訳の科目を書き換える根拠にはしない。
- **部門名の指定方法**：子部門名のみでも「親部門：子部門」形式でも指定可能。ただし同名の子部門が複数ある場合（「その他」など）は必ず「親部門：子部門」形式で指定する。例: 「本店：その他」「CSS：その他」「本店：CA」。一意な子部門名（「AIスカウト」等）はそのまま指定してよい。
- **取引先名は必ずマスタ一覧から正確な名前を使用する**：上記「取引先」マスタの値をそのまま使用する。誤った取引先名を使わない。
- **明細行の部門は現在の値を必ず引き継ぐ**：明細行の内容を変更する場合、変更しない行の部門名・勘定科目・金額は現在値をそのまま維持する。部門を変更しない場合はsection_nameを省略する（システムが自動で現在値を引き継ぐ）。
- **変更内容を具体的に提示**：「仕訳ID XXXXXの金額を○○円→○○円に変更、勘定科目・部門は現在のまま」のように明示する。

## 3. 削除処理
- 必ずsearch_deals/search_invoicesで対象を検索してからIDを取得する（IDを自分で作らない）
- 検索後、そのまま即座にdelete_dealsを呼び出す。事前にユーザーにテキストで確認を求めない。
- delete_dealsを呼び出す際は、search_dealsの結果をdeals_detailパラメータにそのまま渡す。
- フロントエンドが削除確認ボタンを表示するので、AIは「削除しますか？」とテキストで辰ない。
- ユーザーがボタンを押した後に実際の削除が実行される。

## 4. 照会・集計・分析
- list_dealsで仕訳一覧を取得し、集計・分析・報告ができる
- list_invoicesで請求書一覧を取得できる
- get_dealで特定の仕訳の詳細を確認できる
- 「〇〇の売上合計を教えて」「今月の支出一覧を見せて」などの照会指示に対応する

## 5. 添付ファイルの活用
- ファイル内容が【添付ファイルの内容】として含まれる場合は必ず参照・活用する
- 「画像は読めない」「テキストを教えてほしい」とは言わない
- 複数ファイルの場合は文脈から各ファイルの役割を判断する
  例: スクリーンショット＋PDF → スクリーンショットは「登録形式サンプル」、PDFは「処理対象データ」

## 6. 請求書の扱い
- 請求書の日付はbilling_date（請求日）で管理される（issue_dateはnullが多い）
- 請求書登録はregister_invoiceツールを使用する

# 重要な制約
- テキストで「よろしいですか？」と確認を求めない。登録・更新・削除のツールを即座に呼び出すこと。フロントが確認ボタンを自動表示する。
- register_dealを呼び出すと確認ボタンがフロントに表示される。AIは「登録しました」と言わない
- update_dealを呼び出すと確認ボタンがフロントに表示される。AIは「更新しました」と言わない
- delete_dealsを呼び出すと削除確認ボタンがフロントに表示される。AIは「削除しました」と言わない
- 複数件ある場合も全件まとめて一度に処理する（一件ずつ確認しない）"""

        # Function Callingのツール定義
        tools = [
            {
                "type": "function",
                "function": {
                    "name": "search_deals",
                    "description": "freeeの仕訳（取引）を検索する。削除対象の確認や存在確認に使用する。",
                    "parameters": {
                        "type": "object",
                        "properties": {
                            "partner_name": {"type": "string", "description": "取引先名（部分一致）"},
                            "start_issue_date": {"type": "string", "description": "発生日開始（YYYY-MM-DD）"},
                            "end_issue_date": {"type": "string", "description": "発生日終了（YYYY-MM-DD）"},
                            "deal_type": {"type": "string", "enum": ["income", "expense"], "description": "取引種別"},
                        },
                    },
                },
            },
            {
                "type": "function",
                "function": {
                    "name": "search_invoices",
                    "description": "freeeの請求書を検索する。削除対象の確認や存在確認に使用する。請求書の日付はbilling_date（請求日）で管理されている。",
                    "parameters": {
                        "type": "object",
                        "properties": {
                            "partner_name": {"type": "string", "description": "取引先名（マスタの正式名を使用すること）"},
                            "start_issue_date": {"type": "string", "description": "請求日（billing_date）の開始日（YYYY-MM-DD）"},
                            "end_issue_date": {"type": "string", "description": "請求日（billing_date）の終了日（YYYY-MM-DD）"},
                        },
                    },
                },
            },
            {
                "type": "function",
                "function": {
                    "name": "register_deal",
                    "description": "freeeに仕訳（取引）を登録する。必須項目（issue_date・partner_name・due_date・section_name）がすべて揃っていることを確認してから呼び出すこと。",
                    "parameters": {
                        "type": "object",
                        "required": ["deal_type", "issue_date", "due_date", "partner_name", "section_name", "details"],
                        "properties": {
                            "deal_type": {"type": "string", "enum": ["income", "expense"]},
                            "issue_date": {"type": "string", "description": "YYYY-MM-DD"},
                            "due_date": {"type": "string", "description": "YYYY-MM-DD（必須）"},
                            "partner_name": {"type": "string", "description": "取引先名（必須）"},
                            "section_name": {"type": "string", "description": "対応部門名（必須）"},
                            "details": {
                                "type": "array",
                                "items": {
                                    "type": "object",
                                    "required": ["account_item_name", "amount", "tax_code"],
                                    "properties": {
                                        "account_item_name": {"type": "string"},
                                        "amount": {"type": "integer", "description": "税抜金額"},
                                        "tax_code": {"type": "integer", "description": "1=課税売上10%, 7=課税仕入10%, 0=不課税"},
                                        "section_name": {"type": "string"},
                                        "tag_names": {"type": "array", "items": {"type": "string"}},
                                        "description": {"type": "string"},
                                    },
                                },
                            },
                        },
                    },
                },
            },
            {
                "type": "function",
                "function": {
                    "name": "delete_deals",
                    "description": "指定したIDの仕訳を削除する。必ず事前にsearch_dealsで対象を検索し、その検索結果に含まれる実際のid値のみを使用すること。実在しないダミーID（1,2,3など）を渡してはいけない。",
                    "parameters": {
                        "type": "object",
                        "required": ["deal_ids", "confirmation_message"],
                        "properties": {
                            "deal_ids": {"type": "array", "items": {"type": "integer"}, "description": "削除する仕訳IDのリスト。search_dealsの結果から取得した実際のid値のみ使用すること。"},
                            "confirmation_message": {"type": "string", "description": "削除対象の説明文。「　を削除します。」という形式で記述すること。決して「削除しました」と過去形にしないこと。例：「株式会社Stellifyの2026年9月以降の仕訳を削除します。」"},
                            "deals_detail": {"type": "array", "description": "search_dealsで取得した仕訳の詳細情報。各要素はid/issue_date/amount/partner_nameを含む。", "items": {"type": "object"}},
                        },
                    },
                },
            },
            {
                "type": "function",
                "function": {
                    "name": "delete_invoices",
                    "description": "指定したIDの請求書を削除する。必ず事前にsearch_invoicesで対象を検索し、その検索結果に含まれる実際のid値のみを使用すること。実在しないダミーID（1,2,3など）を渡してはいけない。",
                    "parameters": {
                        "type": "object",
                        "required": ["invoice_ids", "confirmation_message"],
                        "properties": {
                            "invoice_ids": {"type": "array", "items": {"type": "integer"}, "description": "削除する請求書IDのリスト。search_invoicesの結果から取得した実際のid値のみ使用すること。"},
                            "confirmation_message": {"type": "string", "description": "削除対象の説明文。「　を削除します。」という形式で記述すること。決して「削除しました」と過去形にしないこと。例：「株式会社Stellifyの2026年9月以降の請求書を取消します。」"},
                        },
                    },
                },
            },
            {
                "type": "function",
                "function": {
                    "name": "create_partner",
                    "description": "freeeに新規取引先を作成する。取引先マスタに存在しない取引先への仕訳登録時に使用する。register_dealの前に呼び出すこと。",
                    "parameters": {
                        "type": "object",
                        "required": ["name"],
                        "properties": {
                            "name": {"type": "string", "description": "取引先名（正式名で登録する）"},
                            "shortcut1": {"type": "string", "description": "ショートカット（省略可）"},
                        },
                    },
                },
            },
            {
                "type": "function",
                "function": {
                    "name": "list_deals",
                    "description": "freeeの仕訳一覧を取得する。照会・集計・分析に使用する。合計金額や件数の確認にも使える。",
                    "parameters": {
                        "type": "object",
                        "properties": {
                            "partner_name": {"type": "string", "description": "取引先名（部分一致）"},
                            "start_issue_date": {"type": "string", "description": "発生日開始（YYYY-MM-DD）"},
                            "end_issue_date": {"type": "string", "description": "発生日終了（YYYY-MM-DD）"},
                            "deal_type": {"type": "string", "enum": ["income", "expense"], "description": "取引種別（income=収入, expense=支出）"},
                            "limit": {"type": "integer", "description": "取得件数（最大10、デフォルト100）"},
                        },
                    },
                },
            },
            {
                "type": "function",
                "function": {
                    "name": "get_deal",
                    "description": "freeeの仕訳1件の詳細を取得する。特定の仕訳の内容を確認する際に使用する。",
                    "parameters": {
                        "type": "object",
                        "required": ["deal_id"],
                        "properties": {
                            "deal_id": {"type": "integer", "description": "仕訳ID"},
                        },
                    },
                },
            },
            {
                "type": "function",
                "function": {
                    "name": "execute_delete_deal",
                    "description": "仕訳1件を実際に削除する。必ず事前にユーザーの承認を得てから呼び出すこと。",
                    "parameters": {
                        "type": "object",
                        "required": ["deal_id"],
                        "properties": {
                            "deal_id": {"type": "integer", "description": "削除する仕訳ID"},
                        },
                    },
                },
            },
            {
                "type": "function",
                "function": {
                    "name": "execute_delete_invoice",
                    "description": "請求書1件を実際に取消する。必ず事前にユーザーの承認を得てから呼び出すこと。",
                    "parameters": {
                        "type": "object",
                        "required": ["invoice_id"],
                        "properties": {
                            "invoice_id": {"type": "integer", "description": "取消する請求書ID"},
                        },
                    },
                },
            },
            {
                "type": "function",
                "function": {
                    "name": "list_invoices",
                    "description": "freeeの請求書一覧を取得する。請求書の照会・集計に使用する。",
                    "parameters": {
                        "type": "object",
                        "properties": {
                            "partner_name": {"type": "string", "description": "取引先名（部分一致）"},
                            "start_billing_date": {"type": "string", "description": "請求日開始（YYYY-MM-DD）"},
                            "end_billing_date": {"type": "string", "description": "請求日終了（YYYY-MM-DD）"},
                            "invoice_status": {"type": "string", "description": "ステータス（draft/issue/sent/unsubmitted/submitted）"},
                        },
                    },
                },
            },
            {
                "type": "function",
                "function": {
                    "name": "update_deal",
                    "description": "freeeの既存仕訳（取引）を更新する。必ず事前にsearch_dealsで対象を検索し、get_dealで現在の明細内容を確認してから呼び出すこと。明細行の部門名はマスタの正確な名前を使用すること（推測や略称不可）。変更しない明細行の部門・勘定科目・金額は必ず現在値をそのまま引き継ぐこと。",
                    "parameters": {
                        "type": "object",
                        "required": ["deal_id"],
                        "properties": {
                            "deal_id": {"type": "integer", "description": "更新する仕訳のID"},
                            "issue_date": {"type": "string", "description": "発生日（YYYY-MM-DD）"},
                            "due_date": {"type": "string", "description": "決済期日（YYYY-MM-DD）"},
                            "partner_name": {"type": "string", "description": "取引先名"},
                            "details": {
                                "type": "array",
                                "description": "明細行を変更する場合に指定する。指定した場合は全明細を上書きする。",
                                "items": {
                                    "type": "object",
                                    "required": ["account_item_name", "amount", "tax_code"],
                                    "properties": {
                                        "account_item_name": {"type": "string", "description": "勘定科目名"},
                                        "amount": {"type": "integer", "description": "税抜金額"},
                                        "tax_code": {"type": "integer", "description": "1=課税売上10%, 7=課税仕入10%, 0=不課税"},
                                        "section_name": {"type": "string", "description": "部門名"},
                                        "description": {"type": "string", "description": "摘要"},
                                    },
                                },
                            },
                        },
                    },
                },
            },
            {
                "type": "function",
                "function": {
                    "name": "register_invoice",
                    "description": "freeeに請求書を登録する。請求書登録が必要な取引先への登録に使用する。",
                    "parameters": {
                        "type": "object",
                        "required": ["issue_date", "due_date", "partner_name", "items"],
                        "properties": {
                            "issue_date": {"type": "string", "description": "請求日（YYYY-MM-DD）"},
                            "due_date": {"type": "string", "description": "入金期日（YYYY-MM-DD）"},
                            "partner_name": {"type": "string", "description": "取引先名"},
                            "title": {"type": "string", "description": "請求書タイトル（省略可）"},
                            "memo": {"type": "string", "description": "備考（省略可）"},
                            "items": {
                                "type": "array",
                                "description": "請求明細行",
                                "items": {
                                    "type": "object",
                                    "required": ["name", "unit_price"],
                                    "properties": {
                                        "name": {"type": "string", "description": "品目名"},
                                        "quantity": {"type": "integer", "description": "数量（デフォルト1）"},
                                        "unit_price": {"type": "integer", "description": "単価（税込）"},
                                        "tax_code": {"type": "integer", "description": "税区分（1=課税10%, 0=不課税）"},
                                        "description": {"type": "string", "description": "備考"},
                                    },
                                },
                            },
                        },
                    },
                },
            },
        ]

        # ツール実行関数
        def execute_tool(name, args):
            from datetime import datetime as _dt3

            if name == "search_deals":
                deals = search_deals(**args)
                return _json.dumps([
                    {"id": d["id"], "issue_date": d.get("issue_date"), "partner_name": d.get("partner_name"),
                     "amount": d.get("amount"), "type": d.get("type")}
                    for d in deals
                ], ensure_ascii=False)

            elif name == "search_invoices":
                invoices = search_invoices(**args)
                return _json.dumps([
                    {"id": inv["id"],
                     "billing_date": inv.get("billing_date"),
                     "issue_date": inv.get("issue_date"),
                     "partner_name": inv.get("partner_name"),
                     "total_amount": inv.get("total_amount"),
                     "invoice_number": inv.get("invoice_number")}
                    for inv in invoices
                ], ensure_ascii=False)

            elif name == "list_deals":
                result = list_deals(**args)
                return _json.dumps(result, ensure_ascii=False)

            elif name == "list_invoices":
                result = list_invoices(
                    partner_name=args.get("partner_name"),
                    start_billing_date=args.get("start_billing_date"),
                    end_billing_date=args.get("end_billing_date"),
                    invoice_status=args.get("invoice_status"),
                )
                return _json.dumps(result, ensure_ascii=False)

            elif name == "get_deal":
                deal = get_deal(args["deal_id"])
                return _json.dumps(deal, ensure_ascii=False)

            elif name == "register_deal":
                # 即時登録せず確認待ちとして返す
                deal_type = args.get("deal_type", "expense")
                return _json.dumps({
                    "status": "pending_register",
                    "deal_args": args,
                    "deal_type": deal_type,
                }, ensure_ascii=False)

            elif name == "register_invoice":
                result = register_invoice_agent(
                    issue_date=args["issue_date"],
                    due_date=args["due_date"],
                    partner_name=args["partner_name"],
                    items=args["items"],
                    title=args.get("title", "請求書"),
                    memo=args.get("memo", ""),
                )
                return _json.dumps({"status": "ok", "id": result.get("id")}, ensure_ascii=False)

            elif name == "update_deal":
                deal_id = args.pop("deal_id")
                # 現在の仕訳情報を取得してベースにする
                current = get_deal(deal_id)
                update_fields = {
                    "issue_date": current.get("issue_date"),
                    "due_date": current.get("due_date"),
                    "type": current.get("type"),
                    "partner_id": current.get("partner_id"),
                }
                # 日付・決済期日の更新
                if "issue_date" in args:
                    update_fields["issue_date"] = args["issue_date"]
                if "due_date" in args:
                    update_fields["due_date"] = args["due_date"]
                # 取引先の更新
                if "partner_name" in args:
                    cache = get_master_cache()
                    partners = cache.get("partners", [])
                    pid = resolve_partner_id(args["partner_name"], partners)
                    if pid:
                        update_fields["partner_id"] = pid
                # 明細の更新
                if "details" in args:
                    cache = get_master_cache()
                    account_items = cache.get("account_items", [])
                    sections = cache.get("sections", [])
                    # 親部門を除外（取引に設定できない）: parent_idがNone/0のものは親部門
                    child_sections = [s for s in sections if s.get("parent_id")]
                    current_details = current.get("details", [])
                    new_details = []
                    for idx, d in enumerate(args["details"]):
                        ai_name = d.get("account_item_name", "")
                        ai_id = next((a["id"] for a in account_items if a["name"] == ai_name), None)
                        if not ai_id:
                            ai_id = next((a["id"] for a in account_items if ai_name in a["name"]), None)
                        detail = {
                            "account_item_id": ai_id,
                            "tax_code": d.get("tax_code", 0),
                            "amount": d.get("amount", 0),
                            "description": d.get("description", ""),
                        }
                        # 既存明細行のidを引き継ぐ（更新時に必要）
                        if idx < len(current_details) and current_details[idx].get("id"):
                            detail["id"] = current_details[idx]["id"]
                        # 既存明細行の品目・メモタグ・セグメントタグを引き継ぐ
                        if idx < len(current_details):
                            cur = current_details[idx]
                            if cur.get("item_id"):
                                detail["item_id"] = cur["item_id"]
                            if cur.get("segment_1_tag_id"):
                                detail["segment_1_tag_id"] = cur["segment_1_tag_id"]
                            if cur.get("segment_2_tag_id"):
                                detail["segment_2_tag_id"] = cur["segment_2_tag_id"]
                            if cur.get("segment_3_tag_id"):
                                detail["segment_3_tag_id"] = cur["segment_3_tag_id"]
                            # メモタグ（tags配列からtag_idsへ変換）
                            if cur.get("tags"):
                                detail["tag_ids"] = [t["id"] for t in cur["tags"] if t.get("id")]
                        if d.get("section_name"):
                            # 「親部門：子部門」形式にも対応
                            from freee_client import _find_section_id as _resolve_sec
                            sec_name = d["section_name"]
                            sec_id = _resolve_sec(sec_name, sections)
                            if sec_id:
                                detail["section_id"] = sec_id
                            elif idx < len(current_details) and current_details[idx].get("section_id"):
                                # 部門名が解決できない場合は現在値を維持
                                detail["section_id"] = current_details[idx]["section_id"]
                        else:
                            # section_nameが指定されていない場合は現在の明細行の部門を引き継ぐ
                            if idx < len(current_details) and current_details[idx].get("section_id"):
                                detail["section_id"] = current_details[idx]["section_id"]
                        new_details.append(detail)
                    update_fields["details"] = new_details
                else:
                    # 明細を指定しない場合は現在の明細をそのまま保持（全フィールド引き継ぎ）
                    update_fields["details"] = [
                        {
                            "id": d.get("id"),
                            "account_item_id": d["account_item_id"],
                            "tax_code": d.get("tax_code", 0),
                            "amount": d.get("amount", 0),
                            "description": d.get("description", ""),
                            **({"section_id": d["section_id"]} if d.get("section_id") else {}),
                            **({"item_id": d["item_id"]} if d.get("item_id") else {}),
                            **({"segment_1_tag_id": d["segment_1_tag_id"]} if d.get("segment_1_tag_id") else {}),
                            **({"segment_2_tag_id": d["segment_2_tag_id"]} if d.get("segment_2_tag_id") else {}),
                            **({"segment_3_tag_id": d["segment_3_tag_id"]} if d.get("segment_3_tag_id") else {}),
                            **({"tag_ids": [t["id"] for t in d["tags"] if t.get("id")]} if d.get("tags") else {}),
                        }
                        for d in current.get("details", [])
                    ]
                result = update_deal(deal_id, update_fields)
                return _json.dumps({"status": "ok", "id": deal_id, "message": f"仕訳ID {deal_id} を更新しました"}, ensure_ascii=False)

            elif name == "execute_delete_deal":
                result = execute_delete_deal(args["deal_id"])
                return _json.dumps(result, ensure_ascii=False)

            elif name == "execute_delete_invoice":
                result = execute_delete_invoice(args["invoice_id"])
                return _json.dumps(result, ensure_ascii=False)

            elif name == "delete_deals":
                # AIがsearch_dealsで取得した詳細情報を直接使う
                deal_ids = args["deal_ids"]
                # AIがdeals_detailを渡してきた場合はそれを使用、なければIDのみ
                deals_detail = args.get("deals_detail") or [{"id": did} for did in deal_ids]
                return _json.dumps({
                    "status": "pending_confirmation",
                    "deal_ids": deal_ids,
                    "deals_detail": deals_detail,
                    "message": args.get("confirmation_message", "")
                }, ensure_ascii=False)

            elif name == "delete_invoices":
                # 後方互換性のため残存
                invoice_ids = args["invoice_ids"]
                return _json.dumps({
                    "status": "pending_confirmation",
                    "invoice_ids": invoice_ids,
                    "invoices_detail": [{"id": iid} for iid in invoice_ids],
                    "message": args.get("confirmation_message", "")
                }, ensure_ascii=False)

            elif name == "create_partner":
                partner = create_partner(
                    name=args["name"],
                    shortcut1=args.get("shortcut1", ""),
                )
                return _json.dumps({
                    "status": "ok",
                    "id": partner.get("id"),
                    "name": partner.get("name"),
                    "message": f"取引先「{partner.get('name')}」を新規登録しました（ID: {partner.get('id')}）"
                }, ensure_ascii=False)

            return "{}"

        # エージェントループ（最大15回までツール呼び出しを繰り返す）
        messages = [{"role": "system", "content": system_prompt}]
        for h in history[-10:]:  # 会話履歴を最大10ターン分保持
            messages.append(h)
        messages.append({"role": "user", "content": user_message})

        pending_deletes = {"deal_ids": [], "invoice_ids": [], "message": "", "deals_detail": [], "invoices_detail": []}
        pending_registers = []  # 登録確認待ちリスト
        registered_ids = []
        final_reply = ""

        for loop_count in range(15):  # 最大15回ループ
            # リトライ付きAPI呼び出し（429エラー対策）
            for _retry in range(4):
                resp = req.post(
                    f"{openai_base}/chat/completions",
                    headers={"Authorization": f"Bearer {openai_key}", "Content-Type": "application/json"},
                    json={"model": "gpt-4o", "messages": messages, "tools": tools, "tool_choice": "auto", "temperature": 0.1},
                    timeout=60,
                )
                if resp.status_code == 429 and _retry < 3:
                    import time
                    wait_sec = int(resp.headers.get("Retry-After", 2 ** (_retry + 1)))
                    logger.warning(f"OpenAI 429 rate limit - {wait_sec}秒待機後リトライ ({_retry+1}/3)")
                    time.sleep(wait_sec)
                    continue
                resp.raise_for_status()
                break
            choice = resp.json()["choices"][0]
            msg = choice["message"]
            messages.append(msg)

            # ツール呼び出しがなければ終了
            if not msg.get("tool_calls"):
                final_reply = msg.get("content", "")
                break

            # 各ツールを実行
            for tc in msg["tool_calls"]:
                fn_name = tc["function"]["name"]
                fn_args = _json.loads(tc["function"]["arguments"])
                tool_result = execute_tool(fn_name, fn_args)
                messages.append({
                    "role": "tool",
                    "tool_call_id": tc["id"],
                    "content": tool_result,
                })

                # 削除・登録待機中の場合はフロントに渡す
                result_obj = _json.loads(tool_result)
                if isinstance(result_obj, dict):
                    if result_obj.get("status") == "pending_confirmation":
                        if "deal_ids" in result_obj:
                            pending_deletes["deal_ids"].extend(result_obj["deal_ids"])
                        if "invoice_ids" in result_obj:
                            pending_deletes["invoice_ids"].extend(result_obj["invoice_ids"])
                        if "deals_detail" in result_obj:
                            pending_deletes["deals_detail"].extend(result_obj["deals_detail"])
                        if "invoices_detail" in result_obj:
                            pending_deletes["invoices_detail"].extend(result_obj["invoices_detail"])
                        pending_deletes["message"] = result_obj.get("message", "")

                    # 登録確認待ちの場合
                    if result_obj.get("status") == "pending_register":
                        pending_registers.append({
                            "deal_args": result_obj.get("deal_args", {}),
                            "deal_type": result_obj.get("deal_type", "expense"),
                        })

            # 削除または登録ツールが呼ばれた場合はここでループを打ち切る
            if pending_deletes["deal_ids"] or pending_deletes["invoice_ids"] or pending_registers:
                break

        # 削除確認待ちがある場合はAIの「削除しました」メッセージを上書き
        has_pending_delete = bool(pending_deletes["deal_ids"] or pending_deletes["invoice_ids"])
        if has_pending_delete:
            deal_count = len(pending_deletes["deal_ids"])
            inv_count = len(pending_deletes["invoice_ids"])
            parts = []
            if deal_count:
                parts.append(f"仕訳 {deal_count}件")
            if inv_count:
                parts.append(f"請求書 {inv_count}件")
            final_reply = f"以下の{' と '.join(parts)}が見つかりました。削除してよろしいですか？"

        # 登録確認待ちがある場合はAIの「登録しました」メッセージを上書き
        if pending_registers:
            count = len(pending_registers)
            final_reply = f"以下の内容でfreeeに登録します。内容を確認してください。"

        return jsonify({
            "reply": final_reply,
            "entries": [],
            "delete_actions": [],
            "pending_deletes": pending_deletes if has_pending_delete else None,
            "pending_registers": pending_registers if pending_registers else None,
            "registered_ids": registered_ids,
        })

    except Exception as e:
        logger.exception("AIエージェントエラー")
        return jsonify({"error": str(e)}), 500



@app.route("/api/assistant/search_delete", methods=["POST"])
def api_assistant_search_delete():
    """削除对象の仕訳・請求書を検索してプレビューを返す"""
    from freee_client import search_deals, search_invoices
    data = request.get_json() or {}
    action = data.get("action", {})
    partner_name = action.get("partner_name")
    start_date = action.get("start_date")
    end_date = action.get("end_date")
    deal_type = action.get("deal_type")
    target = action.get("target", "both")

    # 日付範囲のデフォルト（今年度）
    from datetime import date
    today = date.today()
    fiscal_start = f"{today.year}-04-01" if today.month >= 4 else f"{today.year - 1}-04-01"
    fiscal_end = f"{today.year + 1}-03-31" if today.month >= 4 else f"{today.year}-03-31"
    if not start_date:
        start_date = fiscal_start
    if not end_date:
        end_date = fiscal_end

    try:
        result = {"deals": [], "invoices": []}
        if target in ("deal", "both"):
            deals = search_deals(
                partner_name=partner_name,
                start_issue_date=start_date,
                end_issue_date=end_date,
                deal_type=deal_type,
            )
            result["deals"] = [
                {
                    "id": d["id"],
                    "issue_date": d.get("issue_date"),
                    "partner_name": d.get("partner_name"),
                    "amount": d.get("amount"),
                    "type": d.get("type"),
                    "ref_number": d.get("ref_number"),
                }
                for d in deals
            ]
        if target in ("invoice", "both"):
            invoices = search_invoices(
                partner_name=partner_name,
                start_issue_date=start_date,
                end_issue_date=end_date,
            )
            result["invoices"] = [
                {
                    "id": inv["id"],
                    # freee請求書APIは billing_date を使用（issue_dateは会計APIのフィールド名）
                    "issue_date": inv.get("billing_date") or inv.get("issue_date"),
                    "partner_name": inv.get("partner_name"),
                    "total_amount": inv.get("total_amount"),
                    "invoice_number": inv.get("invoice_number"),
                    "title": inv.get("subject") or inv.get("title"),
                }
                for inv in invoices
            ]
        return jsonify(result)
    except Exception as e:
        logger.exception("削除対象検索エラー")
        return jsonify({"error": str(e)}), 500


@app.route("/api/assistant/execute_register", methods=["POST"])
def api_assistant_execute_register():
    """AIが提案した仕訳をfreeeに登録する（ユーザー確認後に実行）"""
    from freee_client import create_deal, get_master_cache
    from datetime import datetime as _dt3
    data = request.get_json() or {}
    pending_registers = data.get("pending_registers", [])
    results = {"registered": [], "failed": []}

    cache = get_master_cache()
    for item in pending_registers:
        deal_args = item.get("deal_args", {})
        deal_type = item.get("deal_type", "expense")
        # deal_argsにdeal_typeが含まれている場合は除去
        deal_args.pop("deal_type", None)

        # ===== 必須項目バリデーション =====
        missing = []
        if not deal_args.get("issue_date"):
            missing.append("発生日（issue_date）")
        if not deal_args.get("partner_name"):
            missing.append("取引先（partner_name）")
        if not deal_args.get("due_date"):
            missing.append("決済期日（due_date）")
        # section_nameはトップレベルまたはdetails[0]のいずれかにあればOK
        has_section = bool(deal_args.get("section_name")) or any(
            d.get("section_name") for d in deal_args.get("details", [])
        )
        if not has_section:
            missing.append("対応部門（section_name）")
        if missing:
            return jsonify({
                "status": "error",
                "message": f"必須項目が不足しています: {', '.join(missing)}",
                "missing_fields": missing,
            }), 400
        # ===== バリデーションここまで =====

        try:
            result = create_deal(deal_args, deal_type, cache)
            deal_id = result.get("id")
            assistant_log.insert(0, {
                "source": "assistant",
                "freee_id": deal_id,
                "issue_date": deal_args.get("issue_date", ""),
                "partner_name": deal_args.get("partner_name", ""),
                "deal_type": deal_type,
                "amount": sum(d.get("amount", 0) for d in deal_args.get("details", [])),
                "registered_at": _dt3.now().strftime("%Y-%m-%d %H:%M"),
                "note": "AI登録（確認後）",
            })
            del assistant_log[200:]
            results["registered"].append(deal_id)
        except Exception as e:
            results["failed"].append({"error": str(e), "args": deal_args})

    deal_urls = [f"https://secure.freee.co.jp/deals#deal_id={did}" for did in results["registered"]]
    return jsonify({
        "status": "ok",
        "registered": len(results["registered"]),
        "ids": results["registered"],
        "urls": deal_urls,
        "failed": results["failed"],
    })


@app.route("/api/assistant/execute_delete", methods=["POST"])
def api_assistant_execute_delete():
    """仕訳・請求書を実際に削除する（請求書はcancelを使用）"""
    from freee_client import delete_deal, delete_invoice
    data = request.get_json() or {}
    deal_ids = data.get("deal_ids", [])
    invoice_ids = data.get("invoice_ids", [])

    # 請求書から作成された仕訳は削除不可のため、そのエラーはskippedとして扱う
    # 判定は厳密なキーワードのみ（短い文字列による誤マッチを防ぐ）
    INVOICE_LINKED_DEAL_ERRORS = [
        "freee請求書から作成された取引は削除できません",
        "請求書から作成",
        "存在しないか既に削除された",
        "not found",
    ]

    results = {
        "deleted_deals": [],
        "failed_deals": [],
        "skipped_deals": [],  # 請求書から作成されたため削除不可な仕訳
        "deleted_invoices": [],
        "failed_invoices": [],
    }
    for did in deal_ids:
        try:
            delete_deal(int(did))
            results["deleted_deals"].append(did)
        except Exception as e:
            err_str = str(e)
            # 請求書から作成された仕訳は削除不可（freeeの仕様）→ skipped扱い
            if any(kw in err_str for kw in INVOICE_LINKED_DEAL_ERRORS):
                logger.info(f"仕訳削除スキップ(請求書連携): deal_id={did}, reason={err_str[:100]}")
                results["skipped_deals"].append({"id": did, "reason": "請求書から作成された仕訳は請求書取消後に自動処理されます"})
            else:
                results["failed_deals"].append({"id": did, "error": err_str})
    for iid in invoice_ids:
        try:
            delete_invoice(int(iid))
            results["deleted_invoices"].append(iid)
        except Exception as e:
            results["failed_invoices"].append({"id": iid, "error": str(e)})

    return jsonify(results)


# ============================================================
# 証桯アップロード
# ============================================================
@app.route("/api/assistant/upload_receipt", methods=["POST"])
def api_upload_receipt():
    """証憑ファイルをfreeeにアップロードして取引に紐付ける"""
    if 'file' not in request.files:
        return jsonify({"status": "error", "message": "ファイルが指定されていません"}), 400
    file = request.files['file']
    deal_id = request.form.get('deal_id')
    description = request.form.get('description', file.filename)
    try:
        with tempfile.NamedTemporaryFile(delete=False, suffix=Path(file.filename).suffix) as tmp:
            file.save(tmp.name)
            receipt = upload_receipt(
                tmp.name,
                deal_id=int(deal_id) if deal_id else None,
                description=description,
            )
        return jsonify({"status": "ok", "receipt_id": receipt.get("id")})
    except Exception as e:
        logger.exception("証憑アップロードエラー")
        return jsonify({"status": "error", "message": str(e)}), 500


# ============================================================
# ファイル読み取り（AI仕訳提案用）
# ============================================================
@app.route("/api/assistant/extract_file", methods=["POST"])
def api_extract_file():
    """
    アップロードされたファイルからテキストを抽出してAIに渡す
    対応形式: PDF, 画像(PNG/JPG), Excel(xlsx/xls), CSV
    """
    files = request.files.getlist('files')
    user_message = request.form.get('message', '')
    logger.info(f"extract_file: {len(files)}件のファイルを受信")
    if not files or all(f.filename == '' for f in files):
        return jsonify({"contents": []})

    contents = []
    for file in files:
        filename = file.filename
        suffix = Path(filename).suffix.lower()
        try:
            with tempfile.NamedTemporaryFile(delete=False, suffix=suffix) as tmp:
                file.save(tmp.name)
                tmp_path = tmp.name

            text = _extract_file_text(tmp_path, filename, suffix)
            if text:
                contents.append(f"【{filename}】\n{text}")
        except Exception as e:
            logger.exception(f"ファイル読み取りエラー: {filename}")
            contents.append(f"【{filename}】読み取りエラー: {str(e)}")

    return jsonify({"contents": contents})


def _extract_file_text(path: str, filename: str, suffix: str) -> str:
    """ファイルからテキストを抽出する"""
    import subprocess

    # CSV
    if suffix == '.csv':
        import csv
        rows = []
        with open(path, 'r', encoding='utf-8-sig', errors='replace') as f:
            reader = csv.reader(f)
            for i, row in enumerate(reader):
                rows.append(','.join(row))
                if i > 200:  # 最大200行
                    rows.append('...(以下省略)')
                    break
        return '\n'.join(rows)

    # Excel
    if suffix in ('.xlsx', '.xls'):
        try:
            import openpyxl
            wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
            lines = []
            for sheet in wb.sheetnames[:3]:  # 最大3シート
                ws = wb[sheet]
                lines.append(f'=== シート: {sheet} ===')
                for i, row in enumerate(ws.iter_rows(values_only=True)):
                    if any(c is not None for c in row):
                        lines.append('\t'.join([str(c) if c is not None else '' for c in row]))
                    if i > 300:
                        lines.append('...(以下省略)')
                        break
            return '\n'.join(lines)
        except Exception:
            pass
        # xlsはxlrdで試みる
        try:
            import xlrd
            wb = xlrd.open_workbook(path)
            lines = []
            for sheet in wb.sheets()[:3]:
                lines.append(f'=== シート: {sheet.name} ===')
                for i in range(min(sheet.nrows, 300)):
                    lines.append('\t'.join([str(sheet.cell_value(i, j)) for j in range(sheet.ncols)]))
            return '\n'.join(lines)
        except Exception:
            pass
        return ''

    # PDF
    if suffix == '.pdf':
        try:
            result = subprocess.run(
                ['pdftotext', '-layout', path, '-'],
                capture_output=True, text=True, timeout=30
            )
            text = result.stdout.strip()
            if text:
                return text[:8000]  # 最大8000文字
        except Exception:
            pass
        # pdftotext失敗時はpdf2imageで画像化してOCR
        try:
            from pdf2image import convert_from_path
            images = convert_from_path(path, first_page=1, last_page=3, dpi=150)
            texts = []
            for img in images:
                img_path = path + '_page.png'
                img.save(img_path, 'PNG')
                t = _ocr_image_with_openai(img_path)
                if t:
                    texts.append(t)
            return '\n'.join(texts)[:8000]
        except Exception:
            pass
        return ''

    # 画像（PNG/JPG/WEBP/GIF/BMP）
    if suffix in ('.png', '.jpg', '.jpeg', '.webp', '.gif', '.bmp', '.tiff', '.tif'):
        # webp/gif/bmpはJPEGに変換してからOCR（変換失敗時はそのままOCRを試みる）
        if suffix not in ('.png', '.jpg', '.jpeg'):
            try:
                from PIL import Image as PILImage
                img = PILImage.open(path).convert('RGB')
                converted_path = path + '_converted.jpg'
                img.save(converted_path, 'JPEG', quality=95)
                result = _ocr_image_with_openai(converted_path)
                try:
                    import os as _os
                    _os.remove(converted_path)
                except Exception:
                    pass
                return result
            except Exception as e:
                logger.warning(f"画像変換失敗({suffix}): {e} - webpのままOCRを試みます")
                # Pillowがなくてもwebpをそのまま送信（OpenAI Vision APIはwebpをサポート）
                return _ocr_image_with_openai(path, mime_override='image/webp')
        return _ocr_image_with_openai(path)

    return ''


def _ocr_image_with_openai(image_path: str, mime_override: str = None) -> str:
    """OpenAI Vision APIで画像からテキストを抽出する
    注意: OCRは必ず公式OpenAI APIを直接使用する。
    OPENAI_API_BASEはサンドボックス専用プロキシの場合があり、Vision機能をサポートしない可能性があるため。
    mime_override: 指定した場合はそのmimeタイプを使用（webp等のフォールバック用）
    """
    import base64
    import requests as req
    openai_key = os.environ.get("OPENAI_API_KEY", "")
    # Vision用は常に公式OpenAI APIを使用（プロキシは使わない）
    ocr_api_base = "https://api.openai.com/v1"
    if not openai_key:
        return ''
    try:
        with open(image_path, 'rb') as f:
            b64 = base64.b64encode(f.read()).decode()
        suffix = Path(image_path).suffix.lower().lstrip('.')
        if mime_override:
            mime = mime_override
        elif suffix == 'png':
            mime = 'image/png'
        elif suffix == 'webp':
            mime = 'image/webp'
        else:
            mime = 'image/jpeg'
        resp = req.post(
            f"{ocr_api_base}/chat/completions",
            headers={"Authorization": f"Bearer {openai_key}", "Content-Type": "application/json"},
            json={
                "model": "gpt-4o",
                "messages": [{
                    "role": "user",
                    "content": [
                        {"type": "image_url", "image_url": {"url": f"data:{mime};base64,{b64}", "detail": "high"}},
                        {"type": "text", "text": "This is a Japanese financial document. Extract ALL text exactly as it appears. For tables (such as repayment schedules, invoices, receipts), extract every row and column preserving the structure. Use tab or pipe characters to separate columns. Do not skip any rows. Output in Japanese where Japanese text appears."}
                    ]
                }],
                "max_tokens": 2000,
            },
            timeout=30,
        )
        if resp.status_code == 200:
            return resp.json()["choices"][0]["message"]["content"]
        else:
            logger.warning(f"OCR APIエラー: {resp.status_code} {resp.text[:200]}")
    except Exception as e:
        logger.warning(f"OCRエラー: {e}")
    return ''


# ============================================================
# 支払管理・振込データ生成エンドポイント
# ============================================================

@app.route("/api/payment/preview", methods=["GET"])
def api_payment_preview():
    """
    振込対象・アラート対象の仕訳一覧をプレビューする（FBファイル生成前の確認用）
    フロントエンドが期待する形式:
    {
        "groups": [{"due_date": "2025-07-31", "deals": [{"id": ..., "partner_name": ..., "amount": ..., "description": ...}]}],
        "alerts": [{"partner_name": ..., "due_date": ..., "amount": ..., "reason": ...}]
    }
    """
    try:
        from freee_client import get_payment_deals
        from collections import defaultdict
        result = get_payment_deals(months_back=3, alert_days=10)

        # transfer_targets を due_date × partner_name でグループ化（同じ期日・同じ取引先はまとめる）
        # 構造: {due_date: {partner_name: {amount, deal_ids, section_names}}}
        from collections import OrderedDict
        grouped: dict = OrderedDict()  # due_date -> {partner_name -> entry}
        for t in result.get("transfer_targets", []):
            due = t.get("due_date") or "unknown"
            partner = t.get("partner_name", "")
            if due not in grouped:
                grouped[due] = OrderedDict()
            if partner not in grouped[due]:
                grouped[due][partner] = {
                    "partner_name": partner,
                    "amount": 0,
                    "deal_ids": [],
                    "section_names": set(),
                }
            grouped[due][partner]["amount"] += t.get("amount", 0)
            if t.get("deal_id"):
                grouped[due][partner]["deal_ids"].append(t["deal_id"])
            for sn in t.get("section_names", []):
                grouped[due][partner]["section_names"].add(sn)

        groups = []
        for due_date in sorted(grouped.keys()):
            deals = []
            for partner, entry in grouped[due_date].items():
                deals.append({
                    "id": entry["deal_ids"][0] if len(entry["deal_ids"]) == 1 else None,
                    "deal_ids": entry["deal_ids"],
                    "partner_name": entry["partner_name"],
                    "amount": entry["amount"],
                    "description": ", ".join(sorted(entry["section_names"])) or "",
                    "count": len(entry["deal_ids"]),
                })
            groups.append({
                "due_date": due_date,
                "deals": deals,
            })

        # alert_targets を alerts 形式に変換
        alerts = []
        for a in result.get("alert_targets", []):
            reason_map = {
                "no_receipt": "証憑未添付",
                "bank_missing": "銀行口座未登録",
            }
            alerts.append({
                "partner_name": a.get("partner_name", ""),
                "due_date": a.get("due_date", ""),
                "amount": a.get("amount", 0),
                "reason": reason_map.get(a.get("alert_reason", ""), a.get("alert_reason", "")),
                "deal_id": a.get("deal_id"),
            })

        return jsonify({"groups": groups, "alerts": alerts})
    except Exception as e:
        logger.exception("支払管理プレビューエラー")
        return jsonify({"error": str(e)}), 500


@app.route("/api/payment/generate_fb", methods=["POST"])
def api_payment_generate_fb():
    """
    振込対象仕訳から全銀FBファイルを生成してダウンロードさせる
    """
    import io
    data = request.get_json() or {}
    transfer_date = data.get("transfer_date")  # YYYYMMDD
    deal_ids = data.get("deal_ids")  # 選択した仕訳IDリスト（Noneなら全件）

    if not transfer_date:
        from datetime import datetime as _dt
        transfer_date = _dt.now().strftime("%Y%m%d")

    try:
        from freee_client import get_payment_deals
        result = get_payment_deals(months_back=3, alert_days=10)
        targets = result["transfer_targets"]

        # 選択されたIDのみに絞り込む
        if deal_ids:
            targets = [t for t in targets if t["deal_id"] in deal_ids]

        if not targets:
            return jsonify({"error": "振込対象の仕訳がありません"}), 400

        fb_text, summary = build_fb_file(targets, transfer_date)

        # Shift-JIS（全銀標準）でエンコード
        fb_bytes = fb_text.encode("shift_jis", errors="replace")
        filename = f"sogohurikomi_{transfer_date}.txt"

        # FBファイル生成確定 → 対象仕訳に「振込依頼済」タグを付与
        from freee_client import add_furikomi_tag
        tag_results = {}
        for t in targets:
            did = t["deal_id"]
            ok = add_furikomi_tag(did)
            tag_results[did] = ok
            if not ok:
                logger.warning(f"振込依頼済タグ付与失敗: deal_id={did}")

        tagged_count = sum(1 for v in tag_results.values() if v)
        logger.info(f"振込依頼済タグ付与: {tagged_count}/{len(targets)}件")

        _save_execution_log(
            log_type="payment_fb",
            summary={
                "transfer_date": transfer_date,
                "valid_count": summary["valid_count"],
                "total_amount": summary["total_amount"],
                "skipped_count": len(summary["skipped"]),
                "tagged_count": tagged_count,
            },
            detail={
                "targets": [{"partner_name": t.get("partner_name",""), "amount": t.get("amount",0), "due_date": t.get("due_date","")} for t in targets],
                "skipped": summary["skipped"],
            },
            trigger="manual",
            has_error=False
        )

        return send_file(
            io.BytesIO(fb_bytes),
            mimetype="text/plain; charset=shift_jis",
            as_attachment=True,
            download_name=filename,
        ), 200, {
            "X-Transfer-Count": str(summary["valid_count"]),
            "X-Transfer-Amount": str(summary["total_amount"]),
            "X-Skipped-Count": str(len(summary["skipped"])),
            "X-Tagged-Count": str(tagged_count),
        }
    except Exception as e:
        logger.exception("FBファイル生成エラー")
        _save_execution_log(log_type="payment_fb", summary={"valid_count": 0, "total_amount": 0, "error_message": str(e)}, trigger="manual", has_error=True)
        return jsonify({"error": str(e)}), 500


# ============================================================
# 証憑プレビュー・ダウンロードプロキシ
# ============================================================
@app.route("/api/receipt/<int:receipt_id>/download", methods=["GET"])
def api_receipt_download(receipt_id: int):
    """
    freeeの証憑ファイルをプロキシしてブラウザに返す。
    ?inline=1 の場合はContent-Disposition: inline（プレビュー用）
    デフォルトはattachment（ダウンロード）
    """
    import requests as req
    inline = request.args.get("inline", "0") == "1"
    try:
        from freee_client import get_valid_token, FREEE_API_BASE, FREEE_COMPANY_ID
        token = get_valid_token()
        # まず証憑メタ情報を取得してファイル名を得る
        meta_resp = req.get(
            f"{FREEE_API_BASE}/receipts/{receipt_id}",
            headers={"Authorization": f"Bearer {token}"},
            params={"company_id": FREEE_COMPANY_ID},
            timeout=15,
        )
        if meta_resp.status_code != 200:
            return jsonify({"error": f"証憑メタ取得失敗: {meta_resp.status_code}"}), 404
        receipt_meta = meta_resp.json().get("receipt", {})
        filename = receipt_meta.get("file_name") or receipt_meta.get("description") or f"receipt_{receipt_id}"
        mime_type = receipt_meta.get("mime_type") or "application/octet-stream"

        # ファイル本体をダウンロード
        dl_resp = req.get(
            f"{FREEE_API_BASE}/receipts/{receipt_id}/download",
            headers={"Authorization": f"Bearer {token}"},
            params={"company_id": FREEE_COMPANY_ID},
            timeout=30,
            stream=True,
        )
        if dl_resp.status_code != 200:
            return jsonify({"error": f"証憑ダウンロード失敗: {dl_resp.status_code}"}), 502

        # Content-Typeをfreeeのレスポンスから取得（より正確）
        ct = dl_resp.headers.get("Content-Type", mime_type)
        disposition = "inline" if inline else f'attachment; filename="{filename}"'
        return Response(
            dl_resp.content,
            status=200,
            headers={
                "Content-Type": ct,
                "Content-Disposition": disposition,
            },
        )
    except Exception as e:
        logger.exception("証憑ダウンロードエラー")
        return jsonify({"error": str(e)}), 500


# ============================================================
# 会話履歴API
# ============================================================
@app.route("/api/chat_history/<session_id>", methods=["GET"])
def api_get_chat_history(session_id):
    """session_idの会話履歴を返す"""
    try:
        conn = _get_db()
        rows = conn.execute(
            "SELECT role, content FROM chat_sessions WHERE session_id=? ORDER BY seq ASC",
            (session_id,)
        ).fetchall()
        conn.close()
        messages = [{"role": r[0], "content": r[1]} for r in rows]
        return jsonify({"session_id": session_id, "messages": messages})
    except Exception as e:
        logger.exception("chat_history取得エラー")
        return jsonify({"session_id": session_id, "messages": []})


@app.route("/api/chat_history/<session_id>", methods=["POST"])
def api_save_chat_history(session_id):
    """session_idの会話履歴を保存する（全件上書き）"""
    try:
        data = request.get_json(force=True)
        messages = data.get("messages", [])
        conn = _get_db()
        conn.execute("DELETE FROM chat_sessions WHERE session_id=?", (session_id,))
        for i, msg in enumerate(messages):
            conn.execute(
                "INSERT INTO chat_sessions (session_id, seq, role, content) VALUES (?,?,?,?)",
                (session_id, i, msg.get("role", "user"), msg.get("content", ""))
            )
        conn.commit()
        conn.close()
        return jsonify({"status": "ok", "saved": len(messages)})
    except Exception as e:
        logger.exception("chat_history保存エラー")
        return jsonify({"status": "error", "message": str(e)}), 500


@app.route("/api/chat_history/<session_id>", methods=["DELETE"])
def api_delete_chat_history(session_id):
    """session_idの会話履歴を削除する"""
    try:
        conn = _get_db()
        conn.execute("DELETE FROM chat_sessions WHERE session_id=?", (session_id,))
        conn.commit()
        conn.close()
        return jsonify({"status": "ok"})
    except Exception as e:
        logger.exception("chat_history削除エラー")
        return jsonify({"status": "error", "message": str(e)}), 500


@app.route("/api/chat_sessions", methods=["GET"])
def api_list_chat_sessions():
    """会話セッション一覧を返す（実行ログプルダウン用）
    各セッションの最初のメッセージ日時・件数・最初のユーザー発言を返す。
    """
    try:
        conn = _get_db()
        rows = conn.execute(
            """
            SELECT
                cs.session_id,
                MIN(cs.created_at) AS started_at,
                MAX(cs.created_at) AS last_at,
                COUNT(*) AS msg_count,
                (
                    SELECT content FROM chat_sessions sub
                    WHERE sub.session_id = cs.session_id AND sub.role = 'user'
                    ORDER BY sub.seq ASC
                    LIMIT 1
                ) AS first_user_msg
            FROM chat_sessions cs
            GROUP BY cs.session_id
            ORDER BY started_at DESC
            LIMIT 50
            """
        ).fetchall()
        conn.close()
        sessions = []
        for r in rows:
            sessions.append({
                "session_id": r[0],
                "started_at": r[1],
                "last_at": r[2],
                "msg_count": r[3],
                "first_user_msg": (r[4] or "")[:80],
            })
        return jsonify({"sessions": sessions})
    except Exception as e:
        logger.exception("chat_sessions一覧取得エラー")
        return jsonify({"sessions": []})


# ============================================================
# ルールブックページ
# ============================================================
@app.route("/rules")
def rules_page():
    """ルールブック: 機能ルール（rules.py）+ 変更履歴（GitHub API）"""
    import requests as req
    from rules import RULES

    return render_template(
        "rules.html",
        rules=RULES,
    )


@app.route("/api/rules_html/<tab_id>")
def api_rules_html(tab_id):
    """rules_tabs.htmlの指定タブのHTMLフラグメントを返す（index.htmlのインラインルール表示用）
    rules_tabs.htmlは base.htmlをextendsしない独立ファイルなので、リクエストコンテキスト外でもレンダリング可能。
    """
    from bs4 import BeautifulSoup
    try:
        # rules_tabs.htmlを直接レンダリング（base.htmlに依存しない）
        from rules import RULES
        rendered = render_template("rules_tabs.html", rules=RULES)
        soup = BeautifulSoup(rendered, "html.parser")
        tab_div = soup.find("div", {"id": tab_id})
        if not tab_div:
            return jsonify({"html": "", "error": f"tab_id '{tab_id}' not found"}), 404
        # rules-sectionの内容だけ返す（display:noneなどの属性は除去）
        tab_div["class"] = [c for c in tab_div.get("class", []) if c not in ["rules-section", "active"]]
        del tab_div["id"]
        return jsonify({"html": str(tab_div)})
    except Exception as e:
        logger.exception("rules_html取得エラー")
        return jsonify({"html": "", "error": str(e)}), 500


@app.route("/changelog")
def changelog():
    """変更履歴ページ（リポジトリ全体のコミット履歴）"""
    import requests as req
    history = []
    try:
        github_repo = os.environ.get("GITHUB_REPO", "rikutokosa/notion-freee-auto")
        github_token = os.environ.get("GITHUB_TOKEN", "")
        headers = {"Accept": "application/vnd.github.v3+json"}
        if github_token:
            headers["Authorization"] = f"Bearer {github_token}"
        resp = req.get(
            f"https://api.github.com/repos/{github_repo}/commits",
            headers=headers,
            params={"per_page": 100},
            timeout=10,
        )
        if resp.status_code == 200:
            from datetime import datetime, timezone, timedelta
            JST = timezone(timedelta(hours=9))
            for commit in resp.json():
                raw_date = commit.get("commit", {}).get("author", {}).get("date", "")
                try:
                    dt_utc = datetime.fromisoformat(raw_date.replace("Z", "+00:00"))
                    dt_jst = dt_utc.astimezone(JST)
                    date_str = dt_jst.strftime("%Y-%m-%d %H:%M")
                except Exception:
                    date_str = raw_date[:16]
                msg = commit.get("commit", {}).get("message", "").splitlines()[0]
                sha = commit.get("sha", "")[:7]
                history.append({"date": date_str, "message": msg, "sha": sha})
    except Exception:
        pass

    # GitHub APIが使えない場合はgit logにフォールバック
    if not history:
        try:
            import subprocess
            result = subprocess.run(
                ["git", "log", "--format=%ad|%h|%s", "--date=format:%Y-%m-%d %H:%M"],
                capture_output=True, text=True, cwd=os.path.dirname(os.path.abspath(__file__))
            )
            for line in result.stdout.strip().splitlines():
                parts = line.split("|", 2)
                if len(parts) == 3:
                    history.append({"date": parts[0], "sha": parts[1], "message": parts[2]})
        except Exception:
            pass

    return render_template("changelog.html", history=history)


# ============================================================
# メール通知
# ============================================================
def send_notification_email(subject: str, body: str):
    """
    スケジュール実行結果をSlack Webhookで通知する。
    環境変数:
      SLACK_WEBHOOK_URL : Slack Incoming Webhook URL
    """
    import requests as _requests

    slack_webhook_url = os.environ.get("SLACK_WEBHOOK_URL", "")
    if not slack_webhook_url:
        logger.error("Slack通知失敗: SLACK_WEBHOOK_URLが設定されていません")
        return

    try:
        # Slackのメッセージ本文を整形（コードブロックで読みやすく）
        slack_text = f"*{subject}*\n```{body}```"
        resp = _requests.post(
            slack_webhook_url,
            headers={"Content-Type": "application/json"},
            json={"text": slack_text},
            timeout=15,
        )
        if resp.status_code == 200 and resp.text == "ok":
            logger.info("Slack通知送信完了")
        else:
            logger.warning(f"Slack通知失敗 ({resp.status_code}): {resp.text}")
    except Exception as e:
        logger.error(f"Slack通知例外: {e}")


@app.route("/api/scheduled_run", methods=["POST"])
def scheduled_run():
    """
    毎日昼12時のスケジュール実行エンドポイント。
    1. 自動転記（run_once）を実行
    2. 請求書照合（run_matching）を実行
    3. 結果をメールで通知
    """
    from datetime import datetime, timezone, timedelta
    JST = timezone(timedelta(hours=9))
    now_str = datetime.now(JST).strftime("%Y-%m-%d %H:%M")

    lines = []
    lines.append(f"本店経理自動化システム 日次実行レポート")
    lines.append(f"実行日時: {now_str} JST")
    lines.append("=" * 50)

    has_error = False

    # --- 1. 自動転記 ---
    lines.append("")
    lines.append("▶ 自動転記（Notion → freee）")
    try:
        results = run_once(db_type="all", dry_run=False)
        success = sum(1 for r in results if r.get("status") == "success")
        errors  = sum(1 for r in results if r.get("status") == "error")
        reviews = sum(1 for r in results if r.get("status") == "review")
        skips   = sum(1 for r in results if r.get("status") == "skip")
        total   = len(results)
        if total == 0:
            lines.append("  処理対象レコードなし（経理対応待ちなし）")
        else:
            lines.append(f"  合計: {total}件")
            lines.append(f"  成功: {success}件 / エラー: {errors}件 / 要確認: {reviews}件 / スキップ: {skips}件")
        if success > 0:
            lines.append("  [登録済み仕訳]")
            for r in results:
                if r.get("status") == "success":
                    name = r.get("name", "")
                    deal_ids = []
                    if r.get("sales_id"): deal_ids.append(("売上", r["sales_id"]))
                    if r.get("purchase_id"): deal_ids.append(("仕入", r["purchase_id"]))
                    if r.get("pca_id"): deal_ids.append(("PCA", r["pca_id"]))
                    if deal_ids:
                        links = " / ".join(f"{label}: https://secure.freee.co.jp/deals#deal_id={did}" for label, did in deal_ids)
                        lines.append(f"    - {name}: {links}")
                    elif r.get("invoice_id"):
                        lines.append(f"    - {name}: 請求書ID={r['invoice_id']}")
        if reviews > 0:
            has_error = True
            lines.append("  [要確認詳細]")
            for r in results:
                if r.get("status") == "review":
                    name = r.get("name") or r.get("page_id", "")
                    msg  = r.get("message", "")
                    lines.append(f"    - {name}: {msg}")
        if errors > 0:
            has_error = True
            lines.append("  [エラー詳細]")
            for r in results:
                if r.get("status") == "error":
                    name = r.get("name") or r.get("page_id", "")
                    msg  = r.get("message", "")
                    lines.append(f"    - {name}: {msg}")
        _save_execution_log(
            log_type="auto_transfer",
            summary={"total": total, "success": success, "errors": errors, "reviews": reviews, "skips": skips, "db_type": "all"},
            detail={"results": [{"name": r.get("name",""), "status": r.get("status",""), "message": r.get("message",""), "action": r.get("action",""), "sales_id": r.get("sales_id"), "purchase_id": r.get("purchase_id"), "pca_id": r.get("pca_id"), "invoice_id": r.get("invoice_id")} for r in results]},
            trigger="auto",
            has_error=errors > 0
        )
    except Exception as e:
        has_error = True
        lines.append(f"  実行エラー: {e}")
        _save_execution_log(log_type="auto_transfer", summary={"total": 0, "success": 0, "errors": 1, "error_message": str(e)}, trigger="auto", has_error=True)
        logger.exception("スケジュール自動転記エラー")

    # --- 2. 請求書照合 ---
    lines.append("")
    lines.append("▶ 請求書照合（freeeファイルボックス）")
    try:
        match_result = run_matching(dry_run=False)
        matched   = match_result.get("matched_count", 0)
        unmatched = match_result.get("unmatched_count", 0)
        errs      = match_result.get("errors", [])
        total_r   = match_result.get("total_receipts", 0)
        ai_ocr    = match_result.get("ai_ocr_count", 0)
        if total_r == 0:
            lines.append("  未照合の書類なし")
        else:
            lines.append(f"  対象書類: {total_r}件")
            lines.append(f"  照合成功: {matched}件 / 未照合: {unmatched}件")
            if ai_ocr > 0:
                lines.append(f"  AI-OCR使用: {ai_ocr}件")
            # 照合成功分の仕訳URLを記載
            matched_list = match_result.get("matched", [])
            if matched_list:
                lines.append("  [照合済み仕訳]")
                for m in matched_list[:10]:
                    name = (m.get("receipt_description") or m.get("receipt_name") or m.get("name", ""))[:30]
                    deal_id = m.get("deal_id")
                    if deal_id:
                        lines.append(f"    - {name}: https://secure.freee.co.jp/deals#deal_id={deal_id}")
                if len(matched_list) > 10:
                    lines.append(f"    ... 他{len(matched_list) - 10}件")
        if errs:
            has_error = True
            lines.append(f"  [エラー {len(errs)}件]")
            for e in errs[:5]:
                lines.append(f"    - {e}")
        _save_execution_log(
            log_type="invoice_match",
            summary={"total_receipts": total_r, "matched_count": matched, "unmatched_count": unmatched, "ai_ocr_count": ai_ocr},
            detail={"matched": match_result.get("matched", []), "unmatched": match_result.get("unmatched", []), "errors": errs},
            trigger="auto",
            has_error=len(errs) > 0
        )
    except Exception as e:
        has_error = True
        lines.append(f"  実行エラー: {e}")
        _save_execution_log(log_type="invoice_match", summary={"total_receipts": 0, "matched_count": 0, "unmatched_count": 0, "error_message": str(e)}, trigger="auto", has_error=True)
        logger.exception("スケジュール請求書照合エラー")

    lines.append("")
    lines.append("-" * 50)
    lines.append("本店経理自動化システム")

    body = "\n".join(lines)
    subject_prefix = "⚠️ [要確認] " if has_error else "✅ "
    subject = f"{subject_prefix}本店経理自動化 日次実行レポート {now_str}"

    send_notification_email(subject, body)

    return jsonify({
        "status": "ok",
        "has_error": has_error,
        "report": body,
        "executed_at": now_str,
    })


# ============================================================
# 総合振込 支払期日5日前アラートメール
# ============================================================
@app.route("/api/payment_alert", methods=["POST"])
def scheduled_payment_alert():
    """
    毎日昼12時のスケジュール実行エンドポイント。
    支払期日5日以内に振込データがダウンロードされていない取引を検出し、メール通知する。
    """
    from datetime import datetime, timezone, timedelta
    from freee_client import get_payment_deals
    JST = timezone(timedelta(hours=9))
    now_str = datetime.now(JST).strftime("%Y-%m-%d %H:%M")

    try:
        result = get_payment_deals(months_back=3, alert_days=5)
    except Exception as e:
        logger.exception("総合振込アラート取得エラー")
        return jsonify({"status": "error", "message": str(e)}), 500

    transfer_targets = result.get("transfer_targets", [])
    alert_targets = result.get("alert_targets", [])

    # 支払期日5日以内に迫る振込対象（証憑あり・口座登録済み）で振込データ未ダウンロードのもの
    # ※ 「振込データ未ダウンロード」の判定: 「振込依頼済」タグが付いていない = 未処理
    unprocessed_near_due = [
        t for t in transfer_targets
        if t.get("days_until_due") is not None and t["days_until_due"] <= 5
    ]

    # アラート対象（証憑なし or 口座未登録）で支払期日5日以内
    alert_near_due = [
        t for t in alert_targets
        if t.get("days_until_due") is not None and t["days_until_due"] <= 5
    ]

    # アラートが不要な場合はメール送信しない
    if not unprocessed_near_due and not alert_near_due:
        logger.info("総合振込アラート: 対象なし")
        return jsonify({"status": "ok", "message": "アラート対象なし", "alert_count": 0})

    # メール本文作成
    lines = []
    lines.append("本店経理自動化システム 総合振込アラート")
    lines.append(f"実行日時: {now_str} JST")
    lines.append("=" * 50)
    lines.append("")
    lines.append("支払期日5日以内に迫る取引のうち、振込データのダウンロードがまだ行われていないものがあります。")
    lines.append("")

    if alert_near_due:
        lines.append(f"⚠️ 要対応アラート: {len(alert_near_due)}件")
        lines.append("-" * 40)
        for t in sorted(alert_near_due, key=lambda x: x.get("days_until_due", 99)):
            due = t.get("due_date", "-")
            days = t.get("days_until_due", "-")
            partner = t.get("partner_name", "-")
            amount = t.get("amount", 0)
            sections = ", ".join(t.get("section_names", [])) or "-"
            deal_id = t.get("deal_id", "-")
            alert_reason = t.get("alert_reason", "")
            lines.append(f"  ・取引先: {partner}")
            if days == 0:
                lines.append(f"    支払期日: {due}（本日期限）")
            elif days < 0:
                lines.append(f"    支払期日: {due}（{abs(days)}日過ぎ）")
            else:
                lines.append(f"    支払期日: {due}（あと{days}日）")
            lines.append(f"    金額: {amount:,}円")
            lines.append(f"    対応部門: {sections}")
            if alert_reason == "no_receipt":
                lines.append("    ⚠️ 証桫未添付")
            elif alert_reason == "bank_missing":
                lines.append("    ⚠️ 銀行口座未登録")
            if deal_id != "-":
                lines.append(f"    仕訳: https://secure.freee.co.jp/deals#deal_id={deal_id}")
            lines.append("")

    if unprocessed_near_due:
        lines.append(f"▶ 振込未処理（証桫あり・口座登録済み）: {len(unprocessed_near_due)}件")
        lines.append("-" * 40)
        for t in sorted(unprocessed_near_due, key=lambda x: x.get("days_until_due", 99)):
            due = t.get("due_date", "-")
            days = t.get("days_until_due", "-")
            partner = t.get("partner_name", "-")
            amount = t.get("amount", 0)
            sections = ", ".join(t.get("section_names", [])) or "-"
            deal_id = t.get("deal_id", "-")
            lines.append(f"  ・取引先: {partner}")
            if days == 0:
                lines.append(f"    支払期日: {due}（本日期限）")
            elif days < 0:
                lines.append(f"    支払期日: {due}（{abs(days)}日過ぎ）")
            else:
                lines.append(f"    支払期日: {due}（あと{days}日）")
            lines.append(f"    金額: {amount:,}円")
            lines.append(f"    対応部門: {sections}")
            lines.append("")

    lines.append("-" * 50)
    lines.append("本店経理自動化システム")

    body = "\n".join(lines)
    total_alert = len(unprocessed_near_due) + len(alert_near_due)
    subject = f"⚠️ [総合振込アラート] 支払期日5日以内の未処理取引が{total_alert}件あります {now_str}"

    send_notification_email(subject, body)
    logger.info(f"総合振込アラートメール送信完了: {total_alert}件")

    return jsonify({
        "status": "ok",
        "alert_count": total_alert,
        "unprocessed_near_due": len(unprocessed_near_due),
        "alert_near_due": len(alert_near_due),
        "executed_at": now_str,
    })


# ============================================================
# ヘルスチェック（外部サービスマスタとコードの整合性検証）
# ============================================================

@app.route("/api/healthcheck", methods=["GET"])
def api_healthcheck():
    """
    freee・ Notionのマスタ情報とコード内のハードコード値の整合性を検証する。
    不整合がある場合は warnings に詳細を返す。
    """
    warnings = []
    checks = {"freee_sections": "unchecked", "freee_tags": "unchecked", "notion_statuses": "unchecked"}

    # --- freee部門IDの検証 ---
    try:
        from freee_client import get_valid_token, FREEE_API_BASE, FREEE_COMPANY_ID, CSS_SECTION_IDS
        import requests as req
        token = get_valid_token()
        resp = req.get(
            f"{FREEE_API_BASE}/sections",
            headers={"Authorization": f"Bearer {token}"},
            params={"company_id": FREEE_COMPANY_ID},
            timeout=15,
        )
        if resp.status_code == 200:
            sections = resp.json().get("sections", [])
            section_ids_in_freee = {s["id"] for s in sections}
            for sid in CSS_SECTION_IDS:
                if sid not in section_ids_in_freee:
                    warnings.append(f"CSS部門ID {sid} がfreeeに存在しません")
            checks["freee_sections"] = "ok" if not any("CSS部門ID" in w for w in warnings) else "warning"
        else:
            checks["freee_sections"] = "error"
            warnings.append(f"freee部門APIエラー: {resp.status_code}")
    except Exception as e:
        checks["freee_sections"] = "error"
        warnings.append(f"freee部門検証失敗: {str(e)}")

    # --- freee振込依頼済タグIDの検証 ---
    try:
        from freee_client import FURIKOMI_TAG_ID
        resp = req.get(
            f"{FREEE_API_BASE}/tags",
            headers={"Authorization": f"Bearer {token}"},
            params={"company_id": FREEE_COMPANY_ID},
            timeout=15,
        )
        if resp.status_code == 200:
            tags = resp.json().get("tags", [])
            tag_ids_in_freee = {t["id"] for t in tags}
            if FURIKOMI_TAG_ID not in tag_ids_in_freee:
                warnings.append(f"振込依頼済タグID {FURIKOMI_TAG_ID} がfreeeに存在しません")
            checks["freee_tags"] = "ok" if FURIKOMI_TAG_ID in tag_ids_in_freee else "warning"
        else:
            checks["freee_tags"] = "error"
            warnings.append(f"freeeタグAPIエラー: {resp.status_code}")
    except Exception as e:
        checks["freee_tags"] = "error"
        warnings.append(f"freeeタグ検証失敗: {str(e)}")

    # --- Notionステータス検証 ---
    try:
        from notion_client import PENDING_STATUSES_HONTEN, PENDING_STATUSES_PCA
        import requests as req
        notion_token = os.environ.get("NOTION_TOKEN", "")
        notion_db_honten = os.environ.get("NOTION_DB_ID_HONTEN", "")
        notion_db_pca = os.environ.get("NOTION_DB_ID_PCA", "")

        notion_warnings = []
        for db_name, db_id, expected_statuses in [
            ("本店CA", notion_db_honten, PENDING_STATUSES_HONTEN),
            ("PCA", notion_db_pca, PENDING_STATUSES_PCA),
        ]:
            if not db_id or not notion_token:
                continue
            resp = req.post(
                f"https://api.notion.com/v1/databases/{db_id}",
                headers={
                    "Authorization": f"Bearer {notion_token}",
                    "Notion-Version": "2022-06-28",
                },
                timeout=15,
            )
            if resp.status_code == 200:
                props = resp.json().get("properties", {})
                # 請求ステータスプロパティの選択肢を取得
                status_prop = props.get("請求ステータス") or props.get("ステータス")
                if status_prop and status_prop.get("type") == "select":
                    options = [o["name"] for o in status_prop.get("select", {}).get("options", [])]
                    for st in expected_statuses:
                        if st not in options:
                            notion_warnings.append(
                                f"{db_name}DB: ステータス「{st}」がNotionの選択肢に存在しません"
                            )

        warnings.extend(notion_warnings)
        checks["notion_statuses"] = "ok" if not notion_warnings else "warning"
    except Exception as e:
        checks["notion_statuses"] = "error"
        warnings.append(f"Notionステータス検証失敗: {str(e)}")

    status = "healthy" if not warnings else "degraded"
    return jsonify({
        "status": status,
        "checks": checks,
        "warnings": warnings,
    }), 200 if status == "healthy" else 200


# ============================================================
# 起動
# ============================================================
if __name__ == "__main__":
    try:
        get_valid_token()
        logger.info("トークン有効。手動実行待機中。")
    except Exception:
        logger.warning("freeeトークン未設定 → /auth/freee から認証してください")

    port = int(os.environ.get("PORT", 5000))
    app.run(host="0.0.0.0", port=port, debug=False)
